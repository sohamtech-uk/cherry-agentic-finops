from __future__ import annotations

import csv
import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "fixtures" / "private_markets"

CASH_ROWS = [
    {
        "transaction_id": "TXN-2026-0904-001",
        "booking_date": "2026-09-04",
        "direction": "CREDIT",
        "amount_gbp": "875000.00",
        "currency": "GBP",
        "counterparty": "Northbridge Family Office",
        "reference": "NCGFIII-CALL-2026-03 / LP-002",
        "description": "Capital contribution - Northbridge",
        "status": "BOOKED",
    },
    {
        "transaction_id": "TXN-2026-0904-002",
        "booking_date": "2026-09-04",
        "direction": "CREDIT",
        "amount_gbp": "1000000.00",
        "currency": "GBP",
        "counterparty": "Horizon Community Foundation",
        "reference": "NCGFIII-CALL-2026-03 / LP-004",
        "description": "Capital contribution - Horizon",
        "status": "BOOKED",
    },
    {
        "transaction_id": "TXN-2026-0905-003",
        "booking_date": "2026-09-05",
        "direction": "CREDIT",
        "amount_gbp": "1249500.00",
        "currency": "GBP",
        "counterparty": "Oakfield Pension Trust",
        "reference": "NCGFIII-CALL-2026-03 / LP-001",
        "description": "Capital contribution - Oakfield",
        "status": "BOOKED",
    },
]


def build_workbook(path: Path) -> None:
    workbook = Workbook()
    commitments = workbook.active
    commitments.title = "LP_Commitments"
    commitments.append(
        [
            "LP_ID",
            "LP_Name",
            "Total_Commitment_GBP",
            "Called_To_Date_Before_Current_GBP",
            "Current_Call_GBP",
            "Remaining_After_Current_Call_GBP",
            "Due_Date",
            "Call_Notice_ID",
            "Call_Status",
        ]
    )
    commitments.append(
        [
            "LP-001",
            "Oakfield Pension Trust",
            5_000_000,
            2_750_000,
            1_250_000,
            1_000_000,
            date(2026, 9, 6),
            "NCGFIII-CALL-2026-03",
            "PARTIAL_RECEIPT",
        ]
    )
    commitments.append(
        [
            "LP-002",
            "Northbridge Family Office",
            3_500_000,
            1_925_000,
            875_000,
            700_000,
            date(2026, 9, 6),
            "NCGFIII-CALL-2026-03",
            "RECEIVED",
        ]
    )
    commitments.append(
        [
            "LP-003",
            "Willow University Endowment",
            2_500_000,
            1_375_000,
            625_000,
            500_000,
            date(2026, 9, 6),
            "NCGFIII-CALL-2026-03",
            "AWAITING",
        ]
    )
    commitments.append(
        [
            "LP-004",
            "Horizon Community Foundation",
            4_000_000,
            2_200_000,
            1_000_000,
            800_000,
            date(2026, 9, 6),
            "NCGFIII-CALL-2026-03",
            "RECEIVED",
        ]
    )

    history = workbook.create_sheet("Capital_Call_History")
    history.append(
        [
            "LP_ID",
            "Call_Notice_ID",
            "Issue_Date",
            "Due_Date",
            "Call_Amount_GBP",
            "Beneficiary",
            "Bank_Name",
            "Sort_Code",
            "Account_Last4",
            "Settlement_Status",
        ]
    )
    history.append(
        [
            "LP-001",
            "NCGFIII-CALL-2026-01",
            date(2026, 2, 1),
            date(2026, 2, 10),
            1_500_000,
            "Cedar Peak Growth Fund III LP",
            "Cedar Demo Bank plc",
            "00-00-00",
            "2381",
            "SETTLED",
        ]
    )
    history.append(
        [
            "LP-001",
            "NCGFIII-CALL-2026-02",
            date(2026, 5, 5),
            date(2026, 5, 15),
            1_250_000,
            "Cedar Peak Growth Fund III LP",
            "Cedar Demo Bank plc",
            "00-00-00",
            "2381",
            "SETTLED",
        ]
    )
    history.append(
        [
            "LP-001",
            "NCGFIII-CALL-2026-03",
            date(2026, 8, 28),
            date(2026, 9, 6),
            1_250_000,
            "Cedar Peak Growth Fund III LP",
            "Harbour Demo Bank plc",
            "99-99-99",
            "9437",
            "PENDING",
        ]
    )

    approved = workbook.create_sheet("Approved_Bank_Details")
    approved.append(
        [
            "Fund_ID",
            "Fund_Name",
            "Beneficiary",
            "Approved_Bank_Name",
            "Approved_Sort_Code",
            "Approved_Account_Last4",
            "Approval_Status",
            "Approved_From",
            "Approved_By",
        ]
    )
    approved.append(
        [
            "FUND-001",
            "Cedar Peak Growth Fund III LP",
            "Cedar Peak Growth Fund III LP",
            "Cedar Demo Bank plc",
            "00-00-00",
            "2381",
            "APPROVED",
            date(2026, 1, 1),
            "Treasury Control Committee",
        ]
    )

    notes = workbook.create_sheet("Scenario_Notes")
    notes.append(["SYNTHETIC DEMO FIXTURE - Cherry FundFlow"])
    notes.append(["Purpose", "Backup data for the Ylookup x Encode hackathon."])
    notes.append(["Important", "All entities, amounts and bank details are fictional."])
    workbook.save(path)


def build_cash_csv(path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(CASH_ROWS[0]))
        writer.writeheader()
        writer.writerows(CASH_ROWS)


def build_cash_json(path: Path) -> None:
    path.write_text(
        json.dumps({"transactions": CASH_ROWS}, indent=2) + "\n",
        encoding="utf-8",
    )


def build_capital_call_json(path: Path) -> None:
    payload = {
        "document_type": "capital_call",
        "fund_name": "Cedar Peak Growth Fund III LP",
        "investor_name": "Oakfield Pension Trust",
        "lp_reference": "LP-001",
        "notice_id": "NCGFIII-CALL-2026-03",
        "issue_date": "2026-08-28",
        "due_date": "2026-09-06",
        "currency": "GBP",
        "total_commitment": 5_000_000,
        "called_before_current": 2_750_000,
        "current_call": 1_250_000,
        "remaining_after_current": 1_000_000,
        "beneficiary": "Cedar Peak Growth Fund III LP",
        "bank_name": "Harbour Demo Bank plc",
        "sort_code": "99-99-99",
        "account_last4": "9437",
        "payment_reference": "NCGFIII-CALL-2026-03 / LP-001",
        "purpose": "Portfolio acquisition funding and management fee / fund expenses",
        "confidence": 100,
        "source": "fixture",
        "warnings": [],
    }
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    build_workbook(OUT / "02_LP_Commitments_and_Controls.xlsx")
    build_cash_csv(OUT / "03_Fund_Bank_Cash_Transactions.csv")
    build_cash_json(OUT / "03_Fund_Bank_Cash_Transactions.json")
    build_capital_call_json(OUT / "capital_call_fixture.json")
    print(f"Synthetic private-markets fixtures written to {OUT}")


if __name__ == "__main__":
    main()
