from __future__ import annotations

import json
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.agent_tools import (
    compare_dates,
    compare_periods,
    detect_exposure_breaches,
    detect_stale_prices,
    detect_unsettled_trades,
    find_entity,
    find_section,
    get_nav_case_iterations,
    get_nav_iteration_metrics,
    identify_ylookup_workbook,
    inspect_workflow,
    investigate_exception,
    prioritise_exceptions,
    query_database,
    read_document,
    reconcile_cash,
    reconcile_expense_allocations,
    reconcile_investor_gl_workbook,
    reconcile_loader_sample_workbook,
    reconcile_positions,
    reconcile_trades,
    run_daily_fund_health_check,
    run_finance_scenario,
    run_nav_quality_review,
    validate_management_fees,
)
from app.nav_review_history import get_nav_review_history_store


def setup_function() -> None:
    get_nav_review_history_store().clear()


def teardown_function() -> None:
    get_nav_review_history_store().clear()


def _save_workbook(workbook: Workbook, path: Path) -> str:
    workbook.save(path)
    return str(path)


def test_identify_ylookup_workbook_recognises_investor_gl(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Investor-Level GL"
    sheet.append(["Legal Entity", "GL Account", "Trans Type", "Investor", "Deal Name"])
    sheet.append(["Fund A", "1000", "Capital", "LP 1", "Deal 1"])
    path = _save_workbook(workbook, tmp_path / "Investor-Level GL.xlsx")

    profile = identify_ylookup_workbook(path)

    assert profile["kind"] == "investor_gl"


def test_identify_ylookup_workbook_missing_path_raises() -> None:
    with pytest.raises(ValueError, match="does not exist"):
        identify_ylookup_workbook("/nonexistent/workbook.xlsx")


def test_reconcile_investor_gl_workbook_profiles_source(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Investor-Level GL"
    sheet.append(["Legal Entity", "GL Account", "Trans Type", "Investor", "Deal Name"])
    sheet.append(["Fund A", "1000", "Capital", "LP 1", "Deal 1"])
    path = _save_workbook(workbook, tmp_path / "Investor-Level GL.xlsx")

    result = reconcile_investor_gl_workbook(path)

    assert result["workflow"] == "investor_gl_to_loader"
    assert result["row_count"] == 1


def test_reconcile_loader_sample_workbook_reports_missing_fields(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Some Other Column"])
    path = _save_workbook(workbook, tmp_path / "loader_sample.xlsx")

    result = reconcile_loader_sample_workbook(path)

    assert result["status"] == "review_required"
    assert result["required_target_fields_present"] is False


def test_query_database_matches_inspect_workflow() -> None:
    workflow = run_finance_scenario("autonomous")
    workflow_id = workflow["workflow_id"]

    assert query_database(workflow_id) == inspect_workflow(workflow_id)


def _write_nav_summary(path: Path, **overrides: object) -> str:
    payload = {
        "legal_entity": "Fund X",
        "period_end": "2026-06-30",
        "currency": "USD",
        "total_assets": 5_000_000,
        "total_liabilities": 150_000,
        "reported_equity": 4_850_000,
        "opening_nav": 4_700_000,
        "contributions": 250_000,
        "distributions": 100_000,
        "investment_movement": 0,
        "income": 10_000,
        "expenses": 10_000,
        "fx_movement": 0,
        "closing_nav": 4_850_000,
        "investor_capital": [
            {"investor": "Investor A", "reported_capital": 3_000_000},
            {"investor": "Investor B", "reported_capital": 1_850_000},
        ],
    }
    payload.update(overrides)
    path.write_text(json.dumps(payload))
    return str(path)


def test_run_nav_quality_review_clean_summary_is_ready_to_submit(tmp_path: Path) -> None:
    summary_path = _write_nav_summary(tmp_path / "nav-summary.json")

    report = run_nav_quality_review(summary_path)

    assert report["review"]["action"] == "ready_to_submit"
    assert report["review"]["exceptions_open"] == 0
    assert report["root_causes"] == []
    assert report["case_id"].startswith("NAV-")
    assert report["evidence"]["input_sha256"]["nav_summary"]
    assert report["evidence"]["input_sha256"]["source_ledger"] is None
    assert report["iteration"] == {"round_number": 1, "prior_rounds": 0}


def test_run_nav_quality_review_records_a_new_round_each_call(tmp_path: Path) -> None:
    summary_path = _write_nav_summary(tmp_path / "nav-summary.json")

    first = run_nav_quality_review(summary_path)
    second = run_nav_quality_review(summary_path)

    assert first["iteration"]["round_number"] == 1
    assert second["iteration"]["round_number"] == 2


def test_get_nav_case_iterations_reports_recorded_rounds(tmp_path: Path) -> None:
    summary_path = _write_nav_summary(tmp_path / "nav-summary.json")
    run_nav_quality_review(summary_path)

    result = get_nav_case_iterations("Fund X", "2026-06-30")

    assert result["found"] is True
    assert result["rounds_submitted"] == 1
    assert result["closed"] is True


def test_get_nav_case_iterations_reports_not_found_for_unknown_case() -> None:
    result = get_nav_case_iterations("Unknown Fund", "2026-06-30")

    assert result["found"] is False


def test_get_nav_iteration_metrics_reflects_recorded_reviews(tmp_path: Path) -> None:
    summary_path = _write_nav_summary(tmp_path / "nav-summary.json")
    run_nav_quality_review(summary_path)

    metrics = get_nav_iteration_metrics()

    assert metrics["tracked_cases"] >= 1
    assert metrics["closed_cases"] >= 1


def test_run_daily_fund_health_check_classifies_ready_and_attention_needed(
    tmp_path: Path,
) -> None:
    clean_path = _write_nav_summary(
        tmp_path / "clean.json", legal_entity="Fund Ready", period_end="2026-06-30"
    )
    broken_path = _write_nav_summary(
        tmp_path / "broken.json",
        legal_entity="Fund Broken",
        period_end="2026-06-30",
        reported_equity=4_000_000,
    )
    run_nav_quality_review(clean_path)
    run_nav_quality_review(broken_path)

    report = run_daily_fund_health_check()

    assert report["tracked_funds"] == 2
    assert report["ready"] == 1
    assert report["attention_needed"] == 1
    assert report["entries"][0]["legal_entity"] == "Fund Broken"
    assert report["entries"][0]["status"] == "attention_needed"


def test_run_nav_quality_review_flags_balance_sheet_mismatch(tmp_path: Path) -> None:
    summary_path = _write_nav_summary(tmp_path / "nav-summary.json", reported_equity=4_000_000)

    report = run_nav_quality_review(summary_path)

    assert report["review"]["action"] == "return_to_administrator"
    assert any(
        finding["code"] == "balance_sheet.footing_mismatch"
        for finding in report["review"]["findings"]
    )
    assert len(report["root_causes"]) == 1
    assert report["root_causes"][0]["category"] == "balance_sheet"
    assert report["root_causes"][0]["impact_amount"] == "850000.00"


def test_run_nav_quality_review_missing_path_raises() -> None:
    with pytest.raises(ValueError, match="does not exist"):
        run_nav_quality_review("/nonexistent/nav-summary.json")


_CURRENT_STATEMENT = """Subsequent Events
No subsequent events occurred after 2026-06-30.
"""

_PRIOR_STATEMENT = """Subsequent Events
Portfolio Company X completed a transaction on 2026-05-17.
"""


def test_read_document_reads_a_local_file(tmp_path: Path) -> None:
    path = tmp_path / "statement.txt"
    path.write_text(_CURRENT_STATEMENT)

    result = read_document(str(path))

    assert result["document"] == "statement.txt"
    assert "No subsequent events" in result["text"]


def test_read_document_missing_path_raises() -> None:
    with pytest.raises(ValueError, match="does not exist"):
        read_document("/nonexistent/statement.txt")


def test_find_section_reads_a_local_file(tmp_path: Path) -> None:
    path = tmp_path / "statement.txt"
    path.write_text(_CURRENT_STATEMENT)

    result = find_section(str(path), "Subsequent Events")

    assert result["found"] is True


def test_find_entity_reads_a_local_file(tmp_path: Path) -> None:
    path = tmp_path / "statement.txt"
    path.write_text(_CURRENT_STATEMENT)

    result = find_entity(str(path), "2026-06-30")

    assert result["occurrences"] == 1


def test_compare_periods_reads_local_files(tmp_path: Path) -> None:
    current_path = tmp_path / "current.txt"
    current_path.write_text(_CURRENT_STATEMENT)
    prior_path = tmp_path / "prior.txt"
    prior_path.write_text(_PRIOR_STATEMENT)

    result = compare_periods(str(current_path), str(prior_path))

    assert result["identical"] is False
    assert result["current_document"] == "current.txt"
    assert result["prior_document"] == "prior.txt"


def test_compare_dates_reads_local_files(tmp_path: Path) -> None:
    current_path = tmp_path / "current.txt"
    current_path.write_text(_CURRENT_STATEMENT)
    prior_path = tmp_path / "prior.txt"
    prior_path.write_text(_PRIOR_STATEMENT)

    result = compare_dates(str(current_path), str(prior_path))

    assert "2026-06-30" in result["dates_only_in_current"]
    assert "2026-05-17" in result["dates_only_in_prior"]


def test_run_nav_quality_review_rejects_wrong_extension(tmp_path: Path) -> None:
    bad_path = tmp_path / "nav-summary.txt"
    bad_path.write_text("{}")

    with pytest.raises(ValueError, match=r"must be a \.json file"):
        run_nav_quality_review(str(bad_path))


def test_run_nav_quality_review_wraps_invalid_summary_json(tmp_path: Path) -> None:
    bad_path = tmp_path / "nav-summary.json"
    bad_path.write_text("not json")

    with pytest.raises(ValueError, match="Invalid administrator NAV summary"):
        run_nav_quality_review(str(bad_path))


def test_run_nav_quality_review_wraps_corrupt_source_ledger(tmp_path: Path) -> None:
    summary_path = _write_nav_summary(tmp_path / "nav-summary.json")
    ledger_path = tmp_path / "source-ledger.xlsx"
    ledger_path.write_bytes(b"not a real xlsx file")

    with pytest.raises(ValueError, match="Invalid source ledger"):
        run_nav_quality_review(summary_path, str(ledger_path))


def test_run_nav_quality_review_treats_blank_optional_paths_as_not_supplied(
    tmp_path: Path,
) -> None:
    summary_path = _write_nav_summary(tmp_path / "nav-summary.json")

    report = run_nav_quality_review(
        summary_path, source_ledger_path="  ", side_letter_rules_path=""
    )

    assert report["ledger_supplied"] is False
    assert report["side_letter_rules_supplied"] is False


def test_run_nav_quality_review_with_ledger_produces_investor_root_cause(tmp_path: Path) -> None:
    summary_path = _write_nav_summary(
        tmp_path / "nav-summary.json",
        investor_capital=[
            {"investor": "Investor A", "reported_capital": 2_000_000},
            {"investor": "Investor B", "reported_capital": 1_850_000},
        ],
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Investor-Level GL"
    header = [None] * 43
    header[1] = "Static Date"
    header[2] = "Static Date"
    header[3] = "Legal Entity"
    header[21] = "Account Type"
    header[22] = "Trans Type"
    header[23] = "GL Date"
    header[24] = "GL Date"
    header[30] = "Legal Entity Currency"
    header[31] = "Amount (Entity Currency)"
    header[35] = "Investor"
    sheet.append(header)

    def _row(account_type: str, amount: float, investor: str | None = None) -> list[object]:
        row: list[object] = [None] * 43
        row[1], row[2] = date(2026, 4, 1), date(2026, 6, 30)
        row[3] = "Fund X"
        row[21], row[22] = account_type, "Movement"
        row[23], row[24] = date(2026, 5, 1), date(2026, 5, 1)
        row[30], row[31], row[35] = "USD", amount, investor
        return row

    sheet.append(_row("Assets", 5_000_000))
    sheet.append(_row("Liabilities", -150_000))
    sheet.append(_row("Capital", -3_000_000, "Investor A"))
    sheet.append(_row("Capital", -1_850_000, "Investor B"))
    ledger_path = tmp_path / "source-ledger.xlsx"
    workbook.save(ledger_path)

    report = run_nav_quality_review(summary_path, str(ledger_path))

    assert report["ledger_supplied"] is True
    investor_groups = [g for g in report["root_causes"] if g["category"] == "investor_capital"]
    assert any(
        g["investor"] == "Investor A" and g["impact_amount"] == "1000000.00"
        for g in investor_groups
    )


def _write_json(path: Path, payload: object) -> str:
    path.write_text(json.dumps(payload))
    return str(path)


def test_reconcile_positions_tool_flags_a_break(tmp_path: Path) -> None:
    internal_path = _write_json(
        tmp_path / "internal-positions.json",
        [{"fund": "Fund X", "security_id": "SEC1", "quantity": 100, "price": 10}],
    )
    external_path = _write_json(
        tmp_path / "external-positions.json",
        [{"fund": "Fund X", "security_id": "SEC1", "quantity": 90, "price": 10}],
    )

    result = reconcile_positions(internal_path, external_path)

    assert result["matched_count"] == 0
    assert result["breaks"][0]["break_type"] == "quantity_mismatch"
    assert result["exceptions"][0]["category"] == "position"


def test_reconcile_positions_tool_wraps_invalid_json(tmp_path: Path) -> None:
    bad_path = tmp_path / "internal-positions.json"
    bad_path.write_text("not json")
    external_path = _write_json(tmp_path / "external-positions.json", [])

    with pytest.raises(ValueError, match="Invalid internal positions"):
        reconcile_positions(str(bad_path), external_path)


def test_reconcile_cash_tool_flags_a_break(tmp_path: Path) -> None:
    internal_path = _write_json(
        tmp_path / "internal-cash.json",
        [{"fund": "Fund X", "account": "ACC1", "currency": "USD", "balance": 1000}],
    )
    external_path = _write_json(
        tmp_path / "external-cash.json",
        [{"fund": "Fund X", "account": "ACC1", "currency": "USD", "balance": 900}],
    )

    result = reconcile_cash(internal_path, external_path)

    assert result["breaks"][0]["break_type"] == "balance_mismatch"


def test_reconcile_trades_tool_flags_a_break(tmp_path: Path) -> None:
    trade = {
        "trade_id": "T1",
        "fund": "Fund X",
        "security_id": "SEC1",
        "side": "buy",
        "quantity": 100,
        "price": 10,
        "trade_date": "2026-06-01",
    }
    internal_path = _write_json(tmp_path / "internal-trades.json", [trade])
    external_path = _write_json(tmp_path / "external-trades.json", [{**trade, "side": "sell"}])

    result = reconcile_trades(internal_path, external_path)

    assert result["breaks"][0]["break_type"] == "side_mismatch"


def test_detect_stale_prices_tool_flags_a_high_severity_finding(tmp_path: Path) -> None:
    prices_path = _write_json(
        tmp_path / "prices.json",
        [{"security_id": "SEC1", "price": 10, "price_date": "2026-05-01"}],
    )

    result = detect_stale_prices(prices_path, "2026-06-30", max_age_days=3)

    assert result["findings"][0]["severity"] == "high"
    assert result["exceptions"][0]["category"] == "stale_price"


def test_validate_management_fees_tool_flags_amount_mismatch(tmp_path: Path) -> None:
    rules_path = _write_json(
        tmp_path / "fee-rules.json",
        [{"investor": "Investor A", "fee_rate": "0.0125", "fee_basis": "invested_capital"}],
    )
    fees_path = _write_json(
        tmp_path / "administrator-fees.json",
        [
            {
                "investor": "Investor A",
                "fee_basis": "invested_capital",
                "basis_amount": 1_000_000,
                "reported_fee": 12_800,
            }
        ],
    )

    result = validate_management_fees(rules_path, fees_path)

    assert result["breaks"][0]["break_type"] == "amount_mismatch"
    assert result["exceptions"][0]["category"] == "management_fee"


def test_validate_management_fees_tool_wraps_invalid_json(tmp_path: Path) -> None:
    bad_path = tmp_path / "fee-rules.json"
    bad_path.write_text("not json")
    fees_path = _write_json(tmp_path / "administrator-fees.json", [])

    with pytest.raises(ValueError, match="Invalid fee rules"):
        validate_management_fees(str(bad_path), fees_path)


def test_reconcile_expense_allocations_tool_flags_category_mismatch(tmp_path: Path) -> None:
    expected_path = _write_json(
        tmp_path / "expected-allocations.json",
        [{"expense_id": "EXP1", "amount": 5_000, "expected_category": "management_company"}],
    )
    administrator_path = _write_json(
        tmp_path / "administrator-expenses.json",
        [{"expense_id": "EXP1", "amount": 5_000, "allocated_category": "fund"}],
    )

    result = reconcile_expense_allocations(expected_path, administrator_path)

    assert result["breaks"][0]["break_type"] == "category_mismatch"
    assert result["exceptions"][0]["category"] == "expense_allocation"


def test_reconcile_expense_allocations_tool_wraps_invalid_json(tmp_path: Path) -> None:
    bad_path = tmp_path / "expected-allocations.json"
    bad_path.write_text("not json")
    administrator_path = _write_json(tmp_path / "administrator-expenses.json", [])

    with pytest.raises(ValueError, match="Invalid expected expense allocations"):
        reconcile_expense_allocations(str(bad_path), administrator_path)


def test_investigate_exception_tool_defaults_to_highest_priority() -> None:
    exceptions = [
        {
            "category": "trade",
            "code": "trade.price_mismatch",
            "key": "ACC1",
            "title": "Trade break",
            "detail": "Trade break detail.",
            "severity": "high",
            "impact_amount": "100.00",
        },
        {
            "category": "cash",
            "code": "cash.balance_mismatch",
            "key": "ACC1",
            "title": "Cash break",
            "detail": "Cash break detail.",
            "severity": "high",
            "impact_amount": "500.00",
        },
    ]

    result = investigate_exception(exceptions)

    assert result["exception"]["code"] == "cash.balance_mismatch"
    assert [item["code"] for item in result["related_exceptions"]] == ["trade.price_mismatch"]
    assert result["recommended_owner"] == "Treasury / fund controller"
    assert result["next_step"] == "escalate_immediately"


def test_investigate_exception_tool_selects_by_code() -> None:
    exceptions = [
        {
            "category": "stale_price",
            "code": "stale_price.overdue",
            "key": "SEC1",
            "title": "Stale price",
            "detail": "Stale price detail.",
            "severity": "warning",
            "impact_amount": "0",
        }
    ]

    result = investigate_exception(exceptions, code="stale_price.overdue")

    assert result["exception"]["code"] == "stale_price.overdue"
    assert result["next_step"] == "request_evidence"


def test_investigate_exception_tool_raises_when_code_not_found() -> None:
    exceptions = [
        {
            "category": "cash",
            "code": "cash.balance_mismatch",
            "key": "ACC1",
            "title": "Cash break",
            "detail": "Cash break detail.",
            "severity": "high",
            "impact_amount": "500.00",
        }
    ]

    with pytest.raises(ValueError, match="No exception with code"):
        investigate_exception(exceptions, code="does.not.exist")


def test_detect_unsettled_trades_tool_flags_an_overdue_trade(tmp_path: Path) -> None:
    trades_path = _write_json(
        tmp_path / "trades.json",
        [
            {
                "trade_id": "T1",
                "fund": "Fund X",
                "security_id": "SEC1",
                "side": "buy",
                "quantity": 100,
                "price": 10,
                "trade_date": "2026-06-01",
                "settlement_date": "2026-06-01",
                "status": "unsettled",
            }
        ],
    )

    result = detect_unsettled_trades(trades_path, "2026-06-30", grace_days=3)

    assert result["findings"][0]["severity"] == "high"
    assert result["exceptions"][0]["category"] == "unsettled_trade"


def test_detect_exposure_breaches_tool_flags_a_breach(tmp_path: Path) -> None:
    positions_path = _write_json(
        tmp_path / "positions.json",
        [{"fund": "Fund X", "security_id": "SEC1", "quantity": 1, "price": 15}],
    )
    limits_path = _write_json(
        tmp_path / "limits.json",
        [{"label": "Single position cap", "scope": "single_position", "max_percent_of_nav": 10}],
    )

    result = detect_exposure_breaches(positions_path, "100", limits_path)

    assert result["breaches"][0]["key"] == "SEC1"
    assert result["exceptions"][0]["category"] == "exposure_breach"


def test_detect_exposure_breaches_tool_wraps_zero_nav(tmp_path: Path) -> None:
    positions_path = _write_json(tmp_path / "positions.json", [])
    limits_path = _write_json(tmp_path / "limits.json", [])

    with pytest.raises(ValueError, match="Could not compute exposure breaches"):
        detect_exposure_breaches(positions_path, "0", limits_path)


def test_prioritise_exceptions_tool_ranks_by_severity_and_impact() -> None:
    exceptions = [
        {
            "category": "cash",
            "code": "c1",
            "title": "a",
            "detail": "",
            "severity": "warning",
            "impact_amount": 1_000_000,
        },
        {
            "category": "cash",
            "code": "c2",
            "title": "b",
            "detail": "",
            "severity": "high",
            "impact_amount": 10,
        },
    ]

    result = prioritise_exceptions(exceptions)

    assert [item["code"] for item in result["exceptions"]] == ["c2", "c1"]
