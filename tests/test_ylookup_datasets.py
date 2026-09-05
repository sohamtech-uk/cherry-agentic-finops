from __future__ import annotations

import io

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.api import app
from app.ylookup_datasets import analyse_ylookup_dataset_batch, inspect_workbook

client = TestClient(app)


def _xlsx_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _bank_statement_result() -> dict[str, object]:
    workbook = Workbook()
    staging = workbook.active
    staging.title = "Staging Sheet"
    staging.append(
        [
            "Account Name",
            "Account Number",
            "Narrative",
            "Currency",
            "Matched Project Code",
            "Matched Sender/Beneficiary",
            "Classification",
            "Resolved Position",
            "Matched Legal Entity",
        ]
    )
    staging.append(
        [
            "Bank 4319",
            "4319",
            "Payment with unresolved details",
            "EUR",
            "Flag for review - no project match",
            None,
            "Review",
            "Review position",
            "Fund A",
        ]
    )

    diu = workbook.create_sheet("DIU ")
    diu.append(["Batch", "Debit", "Credit"])
    diu.append(["B1", 100, None])
    diu.append(["B1", None, 100])

    positions = workbook.create_sheet("Deal & Position Master List")
    positions.append(["Position"])
    positions.append(["Known Position"])

    account_map = workbook.create_sheet("Account Map")
    account_map.append(["Bank Account"])
    account_map.append(["Operating Account 4319"])

    content = _xlsx_bytes(workbook)
    result = analyse_ylookup_dataset_batch(
        [("Bank statement working file.xlsx", content)],
        ["20260331_FUND_CALDER_EUR_4319.pdf"],
    )
    assert result is not None
    result["evidence"] = {
        "pdf_sha256": [{"file_name": "statement.pdf", "sha256": "a" * 64}],
        "excel_sha256": [{"file_name": "working.xlsx", "sha256": "b" * 64}],
    }
    return result


def test_ylookup_health_declares_native_workflows() -> None:
    response = client.get("/api/ylookup/health")

    assert response.status_code == 200
    body = response.json()
    assert "bank_statements_to_journal_entries" in body["supported_workflows"]
    assert "investor_gl_to_loader" in body["supported_workflows"]
    assert body["input_contract"]["json"] == "not_required"
    assert body["report_formats"] == ["pdf", "xlsx"]


def test_investor_gl_is_not_forced_into_lp_commitments() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Investor-Level GL"
    sheet.append(["Legal Entity", "GL Account", "Trans Type", "Investor", "Deal Name"])
    sheet.append(["Fund A", "1000", "Capital", "LP 1", "Deal 1"])
    content = _xlsx_bytes(workbook)

    profile = inspect_workbook(content, "Investor-Level GL.xlsx")
    result = analyse_ylookup_dataset_batch(
        [("Investor-Level GL.xlsx", content)],
        [],
    )

    assert profile["kind"] == "investor_gl"
    assert result is not None
    workflow = result["workflows"][0]
    assert workflow["workflow"] == "investor_gl_to_loader"
    assert workflow["row_count"] == 1
    assert workflow["status"] == "needs_loader_sample"


def test_bank_statement_workbook_surfaces_actionable_gaps() -> None:
    result = _bank_statement_result()
    workflow = result["workflows"][0]

    assert workflow["workflow"] == "bank_statements_to_journal_entries"
    assert workflow["total_transactions"] == 1
    assert workflow["journal_lines"] == 2
    assert workflow["unmatched_counterparties"] == 1
    assert workflow["project_code_gaps"] == 1
    assert workflow["position_gaps"] == 1
    assert workflow["review_rows"] == 1
    assert workflow["matched_statement_files"] == 1
    assert workflow["review_queue_rows"] == 1
    assert len(workflow["exceptions"]) == 1

    exception = workflow["exceptions"][0]
    assert exception["exception_id"] == "BANK-ROW-2"
    assert exception["account_number"] == "4319"
    assert len(exception["reconciliation_guidance"]) == 4
    owners = {item["owner"] for item in exception["reconciliation_guidance"]}
    assert "Fund operations" in owners
    assert "Fund accounting" in owners
    assert "Investment operations" in owners


def test_ylookup_excel_report_download() -> None:
    result = _bank_statement_result()

    response = client.post("/api/ylookup/report/xlsx", json=result)

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.content.startswith(b"PK")


def test_ylookup_pdf_report_download() -> None:
    result = _bank_statement_result()

    response = client.post("/api/ylookup/report/pdf", json=result)

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/pdf")
    assert response.content.startswith(b"%PDF")
