from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.reconciliation_tools import (
    build_bridge,
    calculate_sum,
    compare_values,
    read_cell,
    read_excel,
)


def _save_workbook(workbook: Workbook, path: Path) -> str:
    workbook.save(path)
    return str(path)


def test_read_excel_returns_headers_and_rows(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Balance Sheet"
    sheet.append(["Account", "Amount"])
    sheet.append(["Assets", 105200000])
    sheet.append(["Liabilities", 12700000])
    path = _save_workbook(workbook, tmp_path / "bs.xlsx")

    result = read_excel(path)

    assert result["sheet_name"] == "Balance Sheet"
    assert result["headers"] == ["Account", "Amount"]
    assert result["row_count_returned"] == 2
    assert result["rows"][0] == ["Assets", 105200000]


def test_read_excel_missing_path_raises() -> None:
    with pytest.raises(ValueError, match="does not exist"):
        read_excel("/nonexistent/workbook.xlsx")


def test_read_cell_returns_value(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Balance Sheet"
    sheet["B12"] = 92500000
    path = _save_workbook(workbook, tmp_path / "bs.xlsx")

    result = read_cell(path, "Balance Sheet", "B12")

    assert result["value"] == 92500000


def test_read_cell_missing_sheet_raises(tmp_path: Path) -> None:
    workbook = Workbook()
    path = _save_workbook(workbook, tmp_path / "bs.xlsx")

    with pytest.raises(ValueError, match="not found"):
        read_cell(path, "Nonexistent Sheet", "A1")


def test_calculate_sum_totals_signed_values() -> None:
    result = calculate_sum(["105200000", "-12700000"])

    assert result["total"] == "92500000.00"


def test_compare_values_matches_within_tolerance() -> None:
    result = compare_values(expected="92500000", actual="92500000.005", tolerance="0.01")

    assert result["matches"] is True
    assert result["status"] == "PASS"


def test_compare_values_fails_outside_tolerance() -> None:
    result = compare_values(expected="92500000", actual="89800000")

    assert result["matches"] is False
    assert result["status"] == "FAIL"
    assert result["difference"] == "2700000.00"


def test_build_bridge_passes_when_it_foots() -> None:
    result = build_bridge(
        opening_balance="2000000",
        movements={"contributions": "500000", "management_fee": "-125000"},
        reported_closing="2375000",
    )

    assert result["status"] == "PASS"
    assert result["expected_closing"] == "2375000.00"


def test_build_bridge_fails_and_reports_difference() -> None:
    result = build_bridge(
        opening_balance="2000000",
        movements={"contributions": "500000", "management_fee": "-125000"},
        reported_closing="2500000",
    )

    assert result["status"] == "FAIL"
    assert result["difference"] == "-125000.00"
