from __future__ import annotations

import csv
import io
import json

import pytest
from openpyxl import Workbook
from reportlab.pdfgen import canvas

from app import fund_manager_classification as fmc
from app.fund_manager_classification import classify_source, classify_sources


def _xlsx_bytes(sheet_titles: list[str]) -> bytes:
    workbook = Workbook()
    workbook.active.title = sheet_titles[0]
    for title in sheet_titles[1:]:
        workbook.create_sheet(title)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _pdf_bytes(*lines: str) -> bytes:
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer)
    for index, line in enumerate(lines):
        pdf.drawString(100, 750 - index * 20, line)
    pdf.save()
    return buffer.getvalue()


def test_classify_source_detects_nav_workbook_by_sheet_name() -> None:
    source = classify_source(_xlsx_bytes(["NAV Summary"]), "Q2_NAV.xlsx", "application/xlsx")

    assert source["detected_type"] == "nav_workbook"
    assert source["status"] == "processed"


def test_classify_source_detects_investor_gl_workbook() -> None:
    source = classify_source(_xlsx_bytes(["Investor-Level GL"]), "gl.xlsx", None)

    assert source["detected_type"] == "investor_gl"


def test_classify_source_falls_back_to_unknown_workbook() -> None:
    source = classify_source(_xlsx_bytes(["Sheet1"]), "mystery.xlsx", None)

    assert source["detected_type"] == "unknown_workbook"
    assert source["status"] == "unknown"


def test_classify_source_detects_capital_call_pdf() -> None:
    source = classify_source(
        _pdf_bytes("Capital Call Notice", "Drawdown Notice for Fund X"), "call.pdf", None
    )

    assert source["detected_type"] == "capital_call_notice"


def test_classify_source_detects_side_letter_pdf() -> None:
    source = classify_source(
        _pdf_bytes("Side Letter Agreement", "Investor: Acme LP"), "sl.pdf", None
    )

    assert source["detected_type"] == "side_letter"


def test_classify_source_falls_back_to_unknown_pdf() -> None:
    source = classify_source(_pdf_bytes("Just some unrelated text"), "mystery.pdf", None)

    assert source["detected_type"] == "unknown_pdf"
    assert source["status"] == "unknown"


def test_classify_source_detects_positions_json() -> None:
    payload = json.dumps([{"security_id": "ABC", "quantity": 100}]).encode()

    source = classify_source(payload, "positions.json", "application/json")

    assert source["detected_type"] == "positions"


def test_classify_source_detects_trades_json_wrapped_in_object() -> None:
    payload = json.dumps({"trades": [{"trade_id": "T1", "side": "buy"}]}).encode()

    source = classify_source(payload, "trades.json", "application/json")

    assert source["detected_type"] == "trades"


def test_classify_source_rejects_invalid_json() -> None:
    source = classify_source(b"{not json", "broken.json", "application/json")

    assert source["detected_type"] == "unknown_json"
    assert source["warnings"]


def test_classify_source_detects_cash_transactions_csv() -> None:
    payload = b"account,currency,balance\nACC-1,USD,1000\n"

    source = classify_source(payload, "cash.csv", "text/csv")

    assert source["detected_type"] == "cash_transactions"


def test_classify_source_empty_file_is_unreadable() -> None:
    source = classify_source(b"", "empty.xlsx", None)

    assert source["status"] == "unreadable"
    assert source["sha256"] is None


def test_classify_source_unsupported_extension_is_unknown() -> None:
    source = classify_source(b"hello", "notes.docx", None)

    assert source["detected_type"] == "unknown"
    assert source["status"] == "unknown"


def test_classify_source_survives_a_workbook_inspection_crash(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A corrupt-but-openable xlsx can fail deep in openpyxl (bad zip member, XML parse error,
    KeyError) rather than at load_workbook() itself, which only raises ValueError. This must be
    classified unknown, not crash the whole batch."""

    def boom(content: bytes, file_name: str) -> dict[str, object]:
        raise KeyError("sheet1.xml")

    monkeypatch.setattr(fmc, "inspect_workbook", boom)

    source = classify_source(b"not-a-real-workbook", "broken.xlsx", None)

    assert source["detected_type"] == "unknown_workbook"
    assert source["status"] == "unknown"
    assert "could not read workbook" in source["warnings"][0]


def test_classify_source_survives_a_document_read_crash(monkeypatch: pytest.MonkeyPatch) -> None:
    def boom(content: bytes, mime_type: str, file_name: str) -> list[tuple[int, str]]:
        raise RuntimeError("native PDF parser crashed")

    monkeypatch.setattr(fmc, "read_document_pages", boom)

    source = classify_source(b"%PDF-1.4 not really a pdf", "broken.pdf", None)

    assert source["detected_type"] == "unknown_pdf"
    assert source["status"] == "unknown"
    assert "could not read document" in source["warnings"][0]


def test_classify_source_survives_a_csv_parse_crash(monkeypatch: pytest.MonkeyPatch) -> None:
    def boom(*args: object, **kwargs: object) -> None:
        raise csv.Error("line contains NUL")

    monkeypatch.setattr(fmc.csv, "DictReader", boom)

    source = classify_source(b"account,currency,balance\nACC-1,USD,1000\n", "broken.csv", None)

    assert source["detected_type"] == "unknown_csv"
    assert source["status"] == "unknown"
    assert "could not parse CSV" in source["warnings"][0]


def test_classify_sources_assigns_sequential_ids() -> None:
    sources = classify_sources(
        [
            ("a.json", json.dumps([{"security_id": "1", "quantity": 1}]).encode(), None),
            ("b.json", json.dumps([{"trade_id": "1", "side": "buy"}]).encode(), None),
        ]
    )

    assert [source["id"] for source in sources] == ["SRC-01", "SRC-02"]
