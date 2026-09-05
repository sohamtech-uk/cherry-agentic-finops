from __future__ import annotations

import io

from openpyxl import Workbook
from reportlab.pdfgen import canvas

from app.private_markets_dataset_analysis import analyse_private_markets_dataset_batch


def _xlsx_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _statement_pdf(account_number: str) -> bytes:
    buffer = io.BytesIO()
    document = canvas.Canvas(buffer)
    document.drawString(72, 760, "Statement details")
    document.drawString(72, 740, "Account number")
    document.drawString(190, 740, account_number)
    document.drawString(72, 720, "Bank name Calder Luxembourg")
    document.save()
    return buffer.getvalue()


def _bank_workbook() -> bytes:
    workbook = Workbook()
    staging = workbook.active
    staging.title = "Staging Sheet"
    staging.append(
        [
            "Account Name",
            "Account Number",
            "Narrative",
            "Currency",
            "Matched Project Code",
            "Matched Sender/Beneficiary",
            "Classification",
            "Resolved Position",
            "Matched Legal Entity",
        ]
    )
    staging.append(
        [
            "NI ABF II SCSP",
            "240-149813-131",
            "Test transaction",
            "DKK",
            "PROJECT-1",
            "Vendor A",
            "Expense",
            "Known Position",
            "Fund A",
        ]
    )

    diu = workbook.create_sheet("DIU ")
    diu.append(["Batch", "Debit", "Credit"])
    diu.append(["B1", 100, None])
    diu.append(["B1", None, 100])

    positions = workbook.create_sheet("Deal & Position Master List")
    positions.append(["Position"])
    positions.append(["Known Position"])

    account_map = workbook.create_sheet("Account Map")
    account_map.append(["Account Number", "Bank Account"])
    account_map.append(["240-149813-131", "NI ABF II - Calder - DKK - 4319"])
    account_map.append(["240-149813-030", "NI ABF II - Calder - EUR - 8102"])
    return _xlsx_bytes(workbook)


def test_pdf_matches_account_map_from_statement_content() -> None:
    result = analyse_private_markets_dataset_batch(
        [("Bank statement to journal entries - working file.xlsx", _bank_workbook())],
        [("downloaded-statement.pdf", _statement_pdf("240-149813-131"))],
    )

    assert result is not None
    workflow = result["workflows"][0]
    assert workflow["matched_statement_files"] == 1
    assert workflow["statement_file_matches"] == [
        {
            "file_name": "downloaded-statement.pdf",
            "match_method": "statement_account_number",
            "account_number": "240149813131",
        }
    ]
    assert "Ylookup" not in result["message"]
    assert "Ylookup" not in workflow["message"]


def test_download_copy_suffix_does_not_break_filename_fallback() -> None:
    result = analyse_private_markets_dataset_batch(
        [("Bank statement to journal entries - working file.xlsx", _bank_workbook())],
        [("20260331_FUND_CALDER_DKK_4319(1).pdf", b"not-extractable-pdf")],
    )

    assert result is not None
    workflow = result["workflows"][0]
    assert workflow["matched_statement_files"] == 1
    assert workflow["statement_file_matches"][0]["match_method"] == "filename_account_code"
    assert workflow["statement_file_matches"][0]["account_code"] == "4319"


def test_workbook_profile_uses_product_neutral_purpose() -> None:
    result = analyse_private_markets_dataset_batch(
        [("Bank statement to journal entries - working file.xlsx", _bank_workbook())],
        [],
    )

    assert result is not None
    assert result["workbook_profiles"][0]["purpose"] == (
        "Bank statements to journal entries working file"
    )
