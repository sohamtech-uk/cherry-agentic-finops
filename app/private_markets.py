from __future__ import annotations

import csv
import io
import json
import logging
from datetime import date
from decimal import ROUND_HALF_UP, Decimal
from enum import StrEnum
from typing import TYPE_CHECKING, Any, Literal

from openpyxl import load_workbook
from pydantic import BaseModel, Field, field_validator, model_validator

if TYPE_CHECKING:
    from app.config import Settings

logger = logging.getLogger(__name__)
TWOPLACES = Decimal("0.01")


def money(value: Decimal | float | int | str | None) -> Decimal:
    return Decimal(str(value or 0)).quantize(TWOPLACES, rounding=ROUND_HALF_UP)


def normalise_last4(value: str | int | None) -> str | None:
    if value is None:
        return None
    digits = "".join(character for character in str(value) if character.isdigit())
    return digits[-4:] if digits else None


def _text(value: Any) -> str:
    return str(value or "").strip()


def _date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip()[:10])


class PrivateMarketsAction(StrEnum):
    AUTO_RECONCILE = "auto_reconcile"
    REQUIRE_APPROVAL = "require_approval"
    REQUEST_EVIDENCE = "request_evidence"


class FindingSeverity(StrEnum):
    PASS = "pass"
    INFO = "info"
    WARNING = "warning"
    HIGH = "high"


class CapitalCallExtraction(BaseModel):
    document_type: Literal["capital_call", "distribution_notice", "unknown"] = "capital_call"
    fund_name: str
    investor_name: str | None = None
    lp_reference: str | None = None
    notice_id: str | None = None
    issue_date: date | None = None
    due_date: date | None = None
    currency: str = "GBP"
    total_commitment: Decimal | None = None
    called_before_current: Decimal | None = None
    current_call: Decimal
    remaining_after_current: Decimal | None = None
    beneficiary: str | None = None
    bank_name: str | None = None
    sort_code: str | None = None
    account_last4: str | None = None
    iban: str | None = None
    swift_bic: str | None = None
    payment_reference: str | None = None
    purpose: str | None = None
    confidence: int = Field(default=0, ge=0, le=100)
    source: Literal["gemini", "manual", "fixture"] = "gemini"
    warnings: list[str] = Field(default_factory=list)

    @field_validator(
        "total_commitment",
        "called_before_current",
        "current_call",
        "remaining_after_current",
        mode="before",
    )
    @classmethod
    def normalise_money(cls, value: Any) -> Decimal | None:
        if value is None or value == "":
            return None
        return money(value)

    @field_validator("currency")
    @classmethod
    def normalise_currency(cls, value: str) -> str:
        return value.upper().strip() or "GBP"

    @field_validator("account_last4", mode="before")
    @classmethod
    def keep_last_four(cls, value: Any) -> str | None:
        return normalise_last4(value)

    @model_validator(mode="after")
    def validate_call(self) -> CapitalCallExtraction:
        if self.current_call <= 0:
            raise ValueError("Current capital call must be greater than zero.")
        if self.issue_date and self.due_date and self.due_date < self.issue_date:
            self.warnings.append("Due date precedes the issue date.")
            self.confidence = min(self.confidence, 70)
        return self


class LPCommitment(BaseModel):
    lp_id: str
    lp_name: str
    total_commitment: Decimal
    called_before_current: Decimal
    current_call: Decimal
    remaining_after_current: Decimal | None = None
    due_date: date | None = None
    call_notice_id: str | None = None
    call_status: str | None = None

    @field_validator(
        "total_commitment",
        "called_before_current",
        "current_call",
        "remaining_after_current",
        mode="before",
    )
    @classmethod
    def normalise_money(cls, value: Any) -> Decimal | None:
        if value is None or value == "":
            return None
        return money(value)


class ApprovedBankDetails(BaseModel):
    fund_id: str | None = None
    fund_name: str
    beneficiary: str | None = None
    bank_name: str | None = None
    sort_code: str | None = None
    account_last4: str | None = None
    approval_status: str | None = None

    @field_validator("account_last4", mode="before")
    @classmethod
    def keep_last_four(cls, value: Any) -> str | None:
        return normalise_last4(value)


class FundCashTransaction(BaseModel):
    transaction_id: str
    booking_date: date
    direction: Literal["credit", "debit"]
    amount: Decimal
    currency: str = "GBP"
    counterparty: str | None = None
    reference: str | None = None
    description: str = ""
    status: str | None = None

    @field_validator("amount", mode="before")
    @classmethod
    def normalise_amount(cls, value: Any) -> Decimal:
        return abs(money(value))

    @field_validator("direction", mode="before")
    @classmethod
    def normalise_direction(cls, value: Any) -> str:
        return _text(value).lower()

    @field_validator("currency", mode="before")
    @classmethod
    def normalise_currency(cls, value: Any) -> str:
        return _text(value).upper() or "GBP"


class PrivateMarketsFinding(BaseModel):
    code: str
    severity: FindingSeverity
    title: str
    detail: str
    expected: str | None = None
    observed: str | None = None


class PrivateMarketsAnalysis(BaseModel):
    action: PrivateMarketsAction
    fund_name: str
    investor_name: str | None = None
    notice_id: str | None = None
    expected_amount: Decimal
    received_amount: Decimal
    variance_amount: Decimal
    due_date: date | None = None
    matched_transaction_ids: list[str] = Field(default_factory=list)
    findings: list[PrivateMarketsFinding] = Field(default_factory=list)
    controls_summary: str
    financial_boundary: str = (
        "Decision support and reconciliation only; this service never initiates a payment."
    )


class PrivateMarketsDataset(BaseModel):
    commitments: list[LPCommitment]
    approved_bank_details: list[ApprovedBankDetails] = Field(default_factory=list)


class GeminiPrivateMarketsUnavailable(RuntimeError):
    pass


CAPITAL_CALL_EXTRACTION_INSTRUCTION = """
You are the document-understanding specialist inside Cherry Agent for private-markets fund operations.
Extract only information visible in the supplied capital-call or distribution notice. Never invent a
fund, investor, commitment, due date, payment instruction, account detail, reference or amount. Use
null for unavailable optional fields and add concise warnings when text is ambiguous. Money fields
must be numeric in the document currency. account_last4 must contain only the final four digits if an
account number is visible. Confidence is an integer from 0 to 100. Return only the schema-conformant
structured response. Do not approve or initiate a payment.
""".strip()


class GeminiCapitalCallExtractor:
    def __init__(self, settings: Settings) -> None:
        self._settings = settings

    async def extract(
        self, content: bytes, mime_type: str, filename: str
    ) -> CapitalCallExtraction:
        if not self._settings.google_ready:
            raise GeminiPrivateMarketsUnavailable(
                "Gemini is not configured. Set GOOGLE_CLOUD_PROJECT for Vertex AI or GOOGLE_API_KEY."
            )
        if not content:
            raise ValueError("The uploaded capital-call document is empty.")

        try:
            from google import genai
            from google.genai import types
        except ImportError as exc:  # pragma: no cover
            raise GeminiPrivateMarketsUnavailable(
                "Install google-genai to process capital-call documents."
            ) from exc

        self._settings.configure_google_environment()
        client = genai.Client()
        async_client = client.aio
        try:
            response = await async_client.models.generate_content(
                model=self._settings.gemini_model,
                contents=[
                    types.Part.from_text(
                        text=(
                            f"Extract private-markets operations data from {filename!r}. "
                            "Return only the structured response required by the schema."
                        )
                    ),
                    types.Part.from_bytes(data=content, mime_type=mime_type),
                ],
                config=types.GenerateContentConfig(
                    system_instruction=CAPITAL_CALL_EXTRACTION_INSTRUCTION,
                    response_mime_type="application/json",
                    response_json_schema=CapitalCallExtraction.model_json_schema(),
                ),
            )
        finally:
            await async_client.aclose()

        if not response.text:
            raise ValueError("Gemini returned no capital-call extraction.")
        try:
            payload: dict[str, Any] = json.loads(response.text)
            payload["source"] = "gemini"
            return CapitalCallExtraction.model_validate(payload)
        except (json.JSONDecodeError, ValueError) as exc:
            logger.exception("Gemini returned an invalid private-markets extraction")
            raise ValueError("Gemini returned an invalid private-markets extraction.") from exc


def parse_commitment_workbook(content: bytes) -> PrivateMarketsDataset:
    if not content:
        raise ValueError("The commitment workbook is empty.")
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    required_sheet = "LP_Commitments"
    if required_sheet not in workbook.sheetnames:
        raise ValueError(f"Workbook must contain a {required_sheet!r} sheet.")

    sheet = workbook[required_sheet]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("LP_Commitments contains no rows.")
    header = {_text(value): index for index, value in enumerate(rows[0]) if value is not None}

    def value(row: tuple[Any, ...], name: str) -> Any:
        index = header.get(name)
        return row[index] if index is not None and index < len(row) else None

    commitments: list[LPCommitment] = []
    for row in rows[1:]:
        if not _text(value(row, "LP_ID")):
            continue
        commitments.append(
            LPCommitment(
                lp_id=_text(value(row, "LP_ID")),
                lp_name=_text(value(row, "LP_Name")),
                total_commitment=value(row, "Total_Commitment_GBP"),
                called_before_current=value(row, "Called_To_Date_Before_Current_GBP"),
                current_call=value(row, "Current_Call_GBP"),
                remaining_after_current=value(row, "Remaining_After_Current_Call_GBP"),
                due_date=_date(value(row, "Due_Date")),
                call_notice_id=_text(value(row, "Call_Notice_ID")) or None,
                call_status=_text(value(row, "Call_Status")) or None,
            )
        )

    approved: list[ApprovedBankDetails] = []
    if "Approved_Bank_Details" in workbook.sheetnames:
        bank_rows = list(workbook["Approved_Bank_Details"].iter_rows(values_only=True))
        if bank_rows:
            bank_header = {
                _text(value): index for index, value in enumerate(bank_rows[0]) if value is not None
            }

            def bank_value(row: tuple[Any, ...], name: str) -> Any:
                index = bank_header.get(name)
                return row[index] if index is not None and index < len(row) else None

            for row in bank_rows[1:]:
                if not _text(bank_value(row, "Fund_Name")):
                    continue
                approved.append(
                    ApprovedBankDetails(
                        fund_id=_text(bank_value(row, "Fund_ID")) or None,
                        fund_name=_text(bank_value(row, "Fund_Name")),
                        beneficiary=_text(bank_value(row, "Beneficiary")) or None,
                        bank_name=_text(bank_value(row, "Approved_Bank_Name")) or None,
                        sort_code=_text(bank_value(row, "Approved_Sort_Code")) or None,
                        account_last4=bank_value(row, "Approved_Account_Last4"),
                        approval_status=_text(bank_value(row, "Approval_Status")) or None,
                    )
                )

    return PrivateMarketsDataset(commitments=commitments, approved_bank_details=approved)


def parse_cash_csv(content: bytes) -> list[FundCashTransaction]:
    if not content:
        raise ValueError("The cash CSV is empty.")
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("The cash CSV must be UTF-8 encoded.") from exc
    reader = csv.DictReader(io.StringIO(text))
    required = {"transaction_id", "booking_date", "direction", "amount_gbp", "currency"}
    if not reader.fieldnames or not required.issubset(set(reader.fieldnames)):
        missing = sorted(required - set(reader.fieldnames or []))
        raise ValueError(f"Cash CSV is missing required columns: {', '.join(missing)}")

    transactions: list[FundCashTransaction] = []
    for row in reader:
        if not _text(row.get("transaction_id")):
            continue
        transactions.append(
            FundCashTransaction(
                transaction_id=_text(row.get("transaction_id")),
                booking_date=_date(row.get("booking_date")),
                direction=_text(row.get("direction")),
                amount=row.get("amount_gbp"),
                currency=_text(row.get("currency")),
                counterparty=_text(row.get("counterparty")) or None,
                reference=_text(row.get("reference")) or None,
                description=_text(row.get("description")),
                status=_text(row.get("status")) or None,
            )
        )
    return transactions


def _find_commitment(
    call: CapitalCallExtraction, commitments: list[LPCommitment]
) -> LPCommitment | None:
    if call.lp_reference:
        for commitment in commitments:
            if commitment.lp_id.casefold() == call.lp_reference.casefold():
                return commitment
    if call.investor_name:
        for commitment in commitments:
            if commitment.lp_name.casefold() == call.investor_name.casefold():
                return commitment
    if call.notice_id:
        candidates = [
            commitment
            for commitment in commitments
            if commitment.call_notice_id
            and commitment.call_notice_id.casefold() == call.notice_id.casefold()
        ]
        if len(candidates) == 1:
            return candidates[0]
    return None


def _matching_cash(
    call: CapitalCallExtraction, transactions: list[FundCashTransaction]
) -> list[FundCashTransaction]:
    matches: list[FundCashTransaction] = []
    notice = (call.notice_id or "").casefold()
    lp_reference = (call.lp_reference or "").casefold()
    investor = (call.investor_name or "").casefold()
    for transaction in transactions:
        if transaction.direction != "credit" or transaction.currency != call.currency:
            continue
        haystack = " ".join(
            filter(
                None,
                [transaction.reference, transaction.description, transaction.counterparty],
            )
        ).casefold()
        if notice and notice in haystack and (not lp_reference or lp_reference in haystack):
            matches.append(transaction)
            continue
        if investor and investor in haystack:
            matches.append(transaction)
    return matches


def analyse_private_markets_case(
    call: CapitalCallExtraction,
    dataset: PrivateMarketsDataset,
    transactions: list[FundCashTransaction],
    *,
    as_of_date: date | None = None,
) -> PrivateMarketsAnalysis:
    findings: list[PrivateMarketsFinding] = []
    commitment = _find_commitment(call, dataset.commitments)

    if commitment is None:
        findings.append(
            PrivateMarketsFinding(
                code="commitment.not_found",
                severity=FindingSeverity.HIGH,
                title="Commitment record not found",
                detail="The capital call could not be tied to a unique LP commitment record.",
            )
        )
    else:
        amount_difference = money(commitment.current_call - call.current_call)
        if amount_difference == 0:
            findings.append(
                PrivateMarketsFinding(
                    code="commitment.call_amount_match",
                    severity=FindingSeverity.PASS,
                    title="Capital call matches commitment schedule",
                    detail="The notice amount matches the current call recorded for the LP.",
                    expected=str(commitment.current_call),
                    observed=str(call.current_call),
                )
            )
        else:
            findings.append(
                PrivateMarketsFinding(
                    code="commitment.call_amount_mismatch",
                    severity=FindingSeverity.HIGH,
                    title="Capital call amount differs from commitment schedule",
                    detail="The PDF call amount and commitment workbook current-call amount differ.",
                    expected=str(commitment.current_call),
                    observed=str(call.current_call),
                )
            )

        calculated_remaining = money(
            commitment.total_commitment
            - commitment.called_before_current
            - commitment.current_call
        )
        if (
            commitment.remaining_after_current is not None
            and calculated_remaining != commitment.remaining_after_current
        ):
            findings.append(
                PrivateMarketsFinding(
                    code="commitment.remaining_math_mismatch",
                    severity=FindingSeverity.HIGH,
                    title="Remaining commitment arithmetic does not reconcile",
                    detail="The workbook remaining commitment differs from deterministic arithmetic.",
                    expected=str(calculated_remaining),
                    observed=str(commitment.remaining_after_current),
                )
            )
        else:
            findings.append(
                PrivateMarketsFinding(
                    code="commitment.remaining_math_valid",
                    severity=FindingSeverity.PASS,
                    title="Remaining commitment arithmetic reconciles",
                    detail="Total commitment less prior calls and the current call reconciles.",
                    observed=str(calculated_remaining),
                )
            )

    approved_bank = next(
        (
            bank
            for bank in dataset.approved_bank_details
            if bank.fund_name.casefold() == call.fund_name.casefold()
            and (bank.approval_status or "APPROVED").upper() == "APPROVED"
        ),
        None,
    )
    if approved_bank and call.account_last4:
        changed_account = (
            approved_bank.account_last4 is not None
            and approved_bank.account_last4 != call.account_last4
        )
        changed_sort = bool(
            approved_bank.sort_code
            and call.sort_code
            and approved_bank.sort_code.replace("-", "") != call.sort_code.replace("-", "")
        )
        if changed_account or changed_sort:
            findings.append(
                PrivateMarketsFinding(
                    code="bank.instructions_changed",
                    severity=FindingSeverity.HIGH,
                    title="Banking instructions changed",
                    detail=(
                        "The notice payment instructions differ from the approved fund banking "
                        "record. Independent verification is required before any payment."
                    ),
                    expected=(
                        f"{approved_bank.bank_name or 'approved bank'} / "
                        f"****{approved_bank.account_last4 or 'unknown'}"
                    ),
                    observed=f"{call.bank_name or 'notice bank'} / ****{call.account_last4}",
                )
            )
        else:
            findings.append(
                PrivateMarketsFinding(
                    code="bank.instructions_match",
                    severity=FindingSeverity.PASS,
                    title="Banking instructions match approved record",
                    detail="The account details on the notice match the approved fund record.",
                )
            )
    elif not approved_bank:
        findings.append(
            PrivateMarketsFinding(
                code="bank.approved_record_missing",
                severity=FindingSeverity.WARNING,
                title="No approved banking record available",
                detail="Bank instructions cannot be independently compared with the supplied data.",
            )
        )

    cash_matches = _matching_cash(call, transactions)
    received = money(sum((transaction.amount for transaction in cash_matches), Decimal("0")))
    variance = money(received - call.current_call)
    if variance == 0:
        findings.append(
            PrivateMarketsFinding(
                code="cash.exact_match",
                severity=FindingSeverity.PASS,
                title="Cash movement exactly matches the capital call",
                detail="Matched credited cash equals the expected call amount.",
                expected=str(call.current_call),
                observed=str(received),
            )
        )
    elif received == 0:
        severity = (
            FindingSeverity.HIGH
            if as_of_date and call.due_date and call.due_date <= as_of_date
            else FindingSeverity.WARNING
        )
        findings.append(
            PrivateMarketsFinding(
                code="cash.missing",
                severity=severity,
                title="No matching cash receipt found",
                detail="No credited cash transaction matches this LP capital call.",
                expected=str(call.current_call),
                observed="0.00",
            )
        )
    elif variance < 0:
        findings.append(
            PrivateMarketsFinding(
                code="cash.short_receipt",
                severity=FindingSeverity.HIGH,
                title="Capital call is under-received",
                detail=f"Matched cash is {abs(variance)} {call.currency} below the call amount.",
                expected=str(call.current_call),
                observed=str(received),
            )
        )
    else:
        findings.append(
            PrivateMarketsFinding(
                code="cash.over_receipt",
                severity=FindingSeverity.HIGH,
                title="Capital call is over-received",
                detail=f"Matched cash exceeds the call amount by {variance} {call.currency}.",
                expected=str(call.current_call),
                observed=str(received),
            )
        )

    high_codes = {finding.code for finding in findings if finding.severity == FindingSeverity.HIGH}
    if "bank.instructions_changed" in high_codes:
        action = PrivateMarketsAction.REQUIRE_APPROVAL
    elif high_codes:
        action = PrivateMarketsAction.REQUEST_EVIDENCE
    elif any(finding.severity == FindingSeverity.WARNING for finding in findings):
        action = PrivateMarketsAction.REQUEST_EVIDENCE
    else:
        action = PrivateMarketsAction.AUTO_RECONCILE

    return PrivateMarketsAnalysis(
        action=action,
        fund_name=call.fund_name,
        investor_name=call.investor_name,
        notice_id=call.notice_id,
        expected_amount=call.current_call,
        received_amount=received,
        variance_amount=variance,
        due_date=call.due_date,
        matched_transaction_ids=[transaction.transaction_id for transaction in cash_matches],
        findings=findings,
        controls_summary=(
            "AI extracts the notice; deterministic commitment, bank-instruction and cash controls "
            "decide whether the case can reconcile, needs approval, or needs more evidence."
        ),
    )
