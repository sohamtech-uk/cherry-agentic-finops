from __future__ import annotations

import io
import os
import re
from typing import Any

from openpyxl import load_workbook


def _text(value: Any) -> str:
    return str(value or "").strip()


def _sheet_by_trimmed_name(workbook: Any, name: str) -> Any | None:
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip() == name:
            return workbook[sheet_name]
    return None


def _first_row(sheet: Any) -> list[Any]:
    try:
        return list(next(sheet.iter_rows(values_only=True)))
    except StopIteration:
        return []


def _header_index(headers: list[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(headers):
        label = _text(value)
        if label and label not in result:
            result[label] = index
    return result


def _cell(row: tuple[Any, ...], indexes: dict[str, int], name: str) -> Any:
    index = indexes.get(name)
    if index is None or index >= len(row):
        return None
    return row[index]


def _guidance_for_reason(reason: str) -> dict[str, Any]:
    if reason == "Counterparty unresolved":
        return {
            "reason": reason,
            "owner": "Fund operations",
            "steps": [
                "Review the bank narrative and reference against the Vendor, Legal Entity and Related Party master lists.",
                "Confirm one canonical sender/beneficiary; do not resolve from amount similarity alone.",
                "Record the selected master-data match in the staging/mapping layer and rerun the control.",
            ],
            "evidence_required": (
                "Master-list record or independently verified supporting document showing the "
                "counterparty identity."
            ),
            "completion_check": (
                "Matched Sender/Beneficiary resolves to exactly one approved master-data record."
            ),
        }
    if reason == "Project code unresolved":
        return {
            "reason": reason,
            "owner": "Fund accounting",
            "steps": [
                "Use the transaction narrative, counterparty and deal context to search the Project Code Report.",
                "Confirm the project belongs to the same legal entity and accounting period.",
                "Apply the approved project-code mapping and rerun journal/control validation.",
            ],
            "evidence_required": "Approved project-code mapping or project accounting support.",
            "completion_check": "Matched Project Code exists in the current Project Code Report.",
        }
    if reason in {"Position not in master list", "Position requires review"}:
        return {
            "reason": reason,
            "owner": "Investment operations",
            "steps": [
                "Compare the proposed position with the Deal & Position Master List using legal entity, deal and instrument context.",
                "If the position is genuinely new, route it through the controlled master-data onboarding process before use.",
                "Update the approved position mapping and rerun the reconciliation control.",
            ],
            "evidence_required": "Approved position master record, deal support or controlled new-position approval.",
            "completion_check": "Resolved Position matches an approved Deal & Position Master List entry.",
        }
    if reason == "Classification flagged Review":
        return {
            "reason": reason,
            "owner": "Fund accounting",
            "steps": [
                "Review transaction type, narrative, counterparty, legal entity and related-party indicators.",
                "Select the accounting classification supported by the source evidence and chart-of-accounts policy.",
                "Regenerate the journal candidate and confirm debit/credit treatment balances.",
            ],
            "evidence_required": "Source transaction support and approved accounting-classification rationale.",
            "completion_check": "Classification is no longer Review and the resulting journal treatment passes controls.",
        }
    return {
        "reason": reason,
        "owner": "Fund operations",
        "steps": [
            "Review the source row and supporting master data.",
            "Record a supported resolution or retain the exception for human review.",
            "Rerun the deterministic controls after the source or mapping is corrected.",
        ],
        "evidence_required": "Evidence supporting the selected reconciliation outcome.",
        "completion_check": "The original exception no longer appears after controls are rerun.",
    }


def inspect_workbook(content: bytes, file_name: str) -> dict[str, Any]:
    """Identify the workbook contract without forcing every XLSX into LP commitments."""

    if not content:
        raise ValueError(f"Workbook {file_name!r} is empty.")
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Workbook {file_name!r} could not be opened: {exc}") from exc

    try:
        sheet_names = list(workbook.sheetnames)
        stripped = {name.strip() for name in sheet_names}
        first_headers: list[str] = []
        if workbook.worksheets:
            first_row = _first_row(workbook.worksheets[0])
            first_headers = [_text(value) for value in first_row if _text(value)]

        kind = "supporting_workbook"
        purpose = "Supporting Excel evidence"
        if "LP_Commitments" in stripped:
            kind = "lp_commitments"
            purpose = "Capital-call commitment and control evidence"
        elif "Staging Sheet" in stripped and "DIU" in stripped:
            kind = "bank_statement_working"
            purpose = "Ylookup bank statements to journal entries working file"
        elif "Investor-Level GL" in stripped:
            kind = "investor_gl"
            purpose = "Ylookup investor-level source general ledger"
        elif {
            "Legal Entity ID",
            "Investor Account ID",
            "Batch Type",
        }.issubset(set(first_headers)):
            kind = "loader_sample"
            purpose = "Ylookup target-system loader sample"
        elif "loader" in file_name.casefold():
            kind = "loader_workbook"
            purpose = "Ylookup loader or mapping workbook"

        return {
            "file_name": file_name,
            "kind": kind,
            "purpose": purpose,
            "sheet_names": sheet_names,
            "used_for_commitment_controls": kind == "lp_commitments",
        }
    finally:
        workbook.close()


def _bank_statement_codes(pdf_file_names: list[str]) -> set[str]:
    codes: set[str] = set()
    for file_name in pdf_file_names:
        stem = os.path.splitext(os.path.basename(file_name))[0]
        candidate = stem.rsplit("_", 1)[-1].strip()
        if candidate.isdigit():
            codes.add(candidate)
    return codes


def analyse_bank_statement_workbook(
    content: bytes,
    file_name: str,
    pdf_file_names: list[str],
) -> dict[str, Any]:
    """Summarise the organiser's bank-statement working file and its review queue."""

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        staging = _sheet_by_trimmed_name(workbook, "Staging Sheet")
        diu = _sheet_by_trimmed_name(workbook, "DIU")
        if staging is None or diu is None:
            message = "Bank-statement workflow requires 'Staging Sheet' and 'DIU' sheets."
            raise ValueError(message)

        headers = _first_row(staging)
        indexes = _header_index(headers)
        required = {
            "Account Name",
            "Narrative",
            "Matched Project Code",
            "Matched Sender/Beneficiary",
            "Classification",
            "Resolved Position",
        }
        missing = sorted(required - set(indexes))
        if missing:
            raise ValueError(
                "Bank-statement staging sheet is missing columns: " + ", ".join(missing)
            )

        master_positions: set[str] = set()
        position_master = _sheet_by_trimmed_name(workbook, "Deal & Position Master List")
        if position_master is not None:
            master_headers = _header_index(_first_row(position_master))
            position_index = master_headers.get("Position")
            if position_index is not None:
                for row in position_master.iter_rows(min_row=2, values_only=True):
                    if position_index < len(row) and _text(row[position_index]):
                        master_positions.add(_text(row[position_index]))

        total_transactions = 0
        unmatched_counterparties = 0
        project_code_gaps = 0
        position_gaps = 0
        review_rows = 0
        exceptions: list[dict[str, Any]] = []

        for row_number, row in enumerate(
            staging.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if not _text(_cell(row, indexes, "Account Name")):
                continue
            total_transactions += 1
            reasons: list[str] = []

            matched_counterparty = _text(_cell(row, indexes, "Matched Sender/Beneficiary"))
            if not matched_counterparty:
                unmatched_counterparties += 1
                reasons.append("Counterparty unresolved")

            matched_project = _text(_cell(row, indexes, "Matched Project Code"))
            if "flag for review" in matched_project.casefold():
                project_code_gaps += 1
                reasons.append("Project code unresolved")

            resolved_position = _text(_cell(row, indexes, "Resolved Position"))
            if resolved_position:
                if resolved_position.casefold().startswith("review"):
                    position_gaps += 1
                    reasons.append("Position requires review")
                elif master_positions and resolved_position not in master_positions:
                    position_gaps += 1
                    reasons.append("Position not in master list")

            classification = _text(_cell(row, indexes, "Classification"))
            if classification.casefold() == "review":
                review_rows += 1
                reasons.append("Classification flagged Review")

            if reasons:
                exceptions.append(
                    {
                        "exception_id": f"BANK-ROW-{row_number}",
                        "row": row_number,
                        "account_name": _text(_cell(row, indexes, "Account Name")),
                        "account_number": _text(_cell(row, indexes, "Account Number")),
                        "currency": _text(_cell(row, indexes, "Currency")),
                        "legal_entity": _text(_cell(row, indexes, "Matched Legal Entity")),
                        "matched_counterparty": matched_counterparty,
                        "matched_project_code": matched_project,
                        "resolved_position": resolved_position,
                        "classification": classification,
                        "related_party_match": _text(
                            _cell(row, indexes, "Related Party Match")
                        ),
                        "reasons": reasons,
                        "narrative": _text(_cell(row, indexes, "Narrative"))[:500],
                        "reconciliation_guidance": [
                            _guidance_for_reason(reason) for reason in reasons
                        ],
                    }
                )

        journal_lines = 0
        for row in diu.iter_rows(min_row=2, values_only=True):
            if any(value is not None and _text(value) for value in row):
                journal_lines += 1

        account_codes: set[str] = set()
        account_map = _sheet_by_trimmed_name(workbook, "Account Map")
        if account_map is not None:
            map_headers = _header_index(_first_row(account_map))
            bank_index = map_headers.get("Bank Account")
            if bank_index is not None:
                for row in account_map.iter_rows(min_row=2, values_only=True):
                    if bank_index >= len(row):
                        continue
                    match = re.search(r"(\d{4,6})\s*$", _text(row[bank_index]))
                    if match:
                        account_codes.add(match.group(1))

        pdf_codes = _bank_statement_codes(pdf_file_names)
        matched_statement_files = len(pdf_codes & account_codes) if account_codes else 0
        journal_expected = total_transactions * 2

        return {
            "workflow": "bank_statements_to_journal_entries",
            "title": "Bank statements → journal entries",
            "status": "review_required" if exceptions else "ready",
            "workbook": file_name,
            "pdf_count": len(pdf_file_names),
            "matched_statement_files": matched_statement_files,
            "total_transactions": total_transactions,
            "journal_lines": journal_lines,
            "journal_expected_lines": journal_expected,
            "journal_line_count_matches": journal_lines == journal_expected,
            "unmatched_counterparties": unmatched_counterparties,
            "project_code_gaps": project_code_gaps,
            "position_gaps": position_gaps,
            "review_rows": review_rows,
            "review_queue_rows": len(exceptions),
            "exceptions": exceptions,
            "sample_exceptions": exceptions[:8],
            "message": (
                "The organiser working file was accepted directly. Review gaps are surfaced as "
                "workflow exceptions rather than treated as invalid input. Click an exception to "
                "see its reconciliation steps and required evidence."
            ),
        }
    finally:
        workbook.close()


def analyse_investor_gl_workbook(content: bytes, file_name: str) -> dict[str, Any]:
    """Profile the organiser's investor-level GL without requiring an LP commitment sheet."""

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        if "Investor-Level GL" not in workbook.sheetnames:
            raise ValueError("Investor GL workflow requires an 'Investor-Level GL' sheet.")
        sheet = workbook["Investor-Level GL"]
        headers = _first_row(sheet)
        indexes = _header_index(headers)
        required = {"Legal Entity", "GL Account", "Trans Type", "Investor"}
        missing = sorted(required - set(indexes))
        if missing:
            raise ValueError("Investor GL is missing columns: " + ", ".join(missing))

        row_count = 0
        legal_entities: set[str] = set()
        fund_families: set[str] = set()
        transaction_currencies: set[str] = set()
        gl_accounts: set[str] = set()
        transaction_types: set[str] = set()
        investors: set[str] = set()
        deals: set[str] = set()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            legal_entity = _text(_cell(row, indexes, "Legal Entity"))
            if not legal_entity:
                continue
            row_count += 1
            legal_entities.add(legal_entity)
            fund_families.add(_text(_cell(row, indexes, "Fund Family")))
            transaction_currencies.add(_text(_cell(row, indexes, "Transaction Currency")))
            gl_accounts.add(_text(_cell(row, indexes, "GL Account")))
            transaction_types.add(_text(_cell(row, indexes, "Trans Type")))
            investors.add(_text(_cell(row, indexes, "Investor")))
            deals.add(_text(_cell(row, indexes, "Deal Name")))

        transaction_currency_count = len({value for value in transaction_currencies if value})
        return {
            "workflow": "investor_gl_to_loader",
            "title": "Investor-level GL → loader",
            "status": "source_profiled",
            "workbook": file_name,
            "row_count": row_count,
            "column_count": len(headers),
            "legal_entity_count": len({value for value in legal_entities if value}),
            "fund_family_count": len({value for value in fund_families if value}),
            "transaction_currency_count": transaction_currency_count,
            "gl_account_count": len({value for value in gl_accounts if value}),
            "transaction_type_count": len({value for value in transaction_types if value}),
            "investor_count": len({value for value in investors if value}),
            "deal_count": len({value for value in deals if value}),
            "message": (
                "The source GL is valid Ylookup evidence. A loader sample or mapping workbook is "
                "needed to complete the source-to-target transformation."
            ),
        }
    finally:
        workbook.close()


def analyse_loader_sample(content: bytes, file_name: str) -> dict[str, Any]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        if not workbook.worksheets:
            raise ValueError("Loader workbook contains no sheets.")
        sheet = workbook.worksheets[0]
        headers = [_text(value) for value in _first_row(sheet) if _text(value)]
        required = {"Legal Entity ID", "Investor Account ID", "Batch Type"}
        return {
            "workflow": "loader_target_contract",
            "title": "Target loader contract",
            "status": "ready" if required.issubset(set(headers)) else "review_required",
            "workbook": file_name,
            "target_column_count": len(headers),
            "required_target_fields_present": required.issubset(set(headers)),
            "target_columns": headers[:40],
        }
    finally:
        workbook.close()


def analyse_ylookup_dataset_batch(
    workbook_items: list[tuple[str, bytes]],
    pdf_file_names: list[str],
) -> dict[str, Any] | None:
    """Auto-route the sponsor datasets instead of treating every workbook as LP commitments."""

    profiles = [inspect_workbook(content, file_name) for file_name, content in workbook_items]
    recognised = {
        "bank_statement_working",
        "investor_gl",
        "loader_sample",
        "loader_workbook",
    }
    if not any(profile["kind"] in recognised for profile in profiles):
        return None

    workflows: list[dict[str, Any]] = []
    loader_analysis: dict[str, Any] | None = None

    for (file_name, content), profile in zip(workbook_items, profiles, strict=True):
        kind = profile["kind"]
        if kind == "bank_statement_working":
            workflows.append(analyse_bank_statement_workbook(content, file_name, pdf_file_names))
        elif kind == "investor_gl":
            workflows.append(analyse_investor_gl_workbook(content, file_name))
        elif kind in {"loader_sample", "loader_workbook"}:
            loader_analysis = analyse_loader_sample(content, file_name)
            workflows.append(loader_analysis)

    if loader_analysis is not None:
        for workflow in workflows:
            if workflow.get("workflow") != "investor_gl_to_loader":
                continue
            workflow["loader_sample_supplied"] = True
            workflow["loader_sample"] = loader_analysis["workbook"]
            workflow["status"] = "ready_for_mapping"
            workflow["message"] = (
                "Source GL and target loader contract are both present. Mapping gaps should be "
                "surfaced explicitly rather than silently filled."
            )
    else:
        for workflow in workflows:
            if workflow.get("workflow") == "investor_gl_to_loader":
                workflow["loader_sample_supplied"] = False
                workflow["status"] = "needs_loader_sample"

    return {
        "workflow_type": "ylookup_dataset_batch",
        "synthetic": False,
        "workbook_profiles": profiles,
        "workflows": workflows,
        "input_summary": {
            "pdf_count": len(pdf_file_names),
            "excel_count": len(workbook_items),
        },
        "message": (
            "Ylookup sponsor workbooks were auto-detected and routed to their native workflows; "
            "no LP_Commitments sheet is required for bank-statement or investor-GL datasets."
        ),
    }
