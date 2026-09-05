from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _text(value: Any) -> str:
    return str(value or "").strip()


def _workflow_exceptions(result: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for workflow in result.get("workflows", []):
        if not isinstance(workflow, dict):
            continue
        exceptions = workflow.get("exceptions") or workflow.get("sample_exceptions") or []
        for item in exceptions:
            if not isinstance(item, dict):
                continue
            rows.append(
                {
                    "workflow": workflow.get("title", workflow.get("workflow", "Workflow")),
                    "workbook": workflow.get("workbook", ""),
                    **item,
                }
            )
    return rows


def _guidance_text(item: dict[str, Any]) -> str:
    parts: list[str] = []
    for section in item.get("reconciliation_guidance", []):
        if not isinstance(section, dict):
            continue
        reason = _text(section.get("reason"))
        steps = "; ".join(_text(step) for step in section.get("steps", []) if _text(step))
        if reason and steps:
            parts.append(f"{reason}: {steps}")
        elif steps:
            parts.append(steps)
    return " | ".join(parts)


def _autosize(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, column in enumerate(sheet.columns, start=1):
            width = min(
                60,
                max(12, max((len(_text(cell.value)) for cell in column), default=0) + 2),
            )
            sheet.column_dimensions[get_column_letter(index)].width = width


def build_ylookup_excel_report(result: dict[str, Any]) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    if summary is None:
        raise RuntimeError("Could not initialise the report workbook.")
    summary.title = "Summary"
    summary.append(["Cherry FundOps — Ylookup Review Report"])
    summary["A1"].font = Font(bold=True, size=16)
    summary.append(["Generated", datetime.now(UTC).isoformat()])
    input_summary = result.get("input_summary", {})
    summary.append(["PDF files", input_summary.get("pdf_count", 0)])
    summary.append(["Excel workbooks", input_summary.get("excel_count", 0)])
    summary.append(["Workflow count", len(result.get("workflows", []))])
    summary.append([])
    summary.append(["Workflow", "Status", "Workbook", "Key metrics"])

    for workflow in result.get("workflows", []):
        if not isinstance(workflow, dict):
            continue
        metric_keys = [
            "total_transactions",
            "journal_lines",
            "review_queue_rows",
            "row_count",
            "investor_count",
            "deal_count",
            "target_column_count",
        ]
        metrics = ", ".join(f"{key}={workflow[key]}" for key in metric_keys if key in workflow)
        summary.append(
            [
                workflow.get("title", workflow.get("workflow", "")),
                workflow.get("status", ""),
                workflow.get("workbook", ""),
                metrics,
            ]
        )

    exceptions_sheet = workbook.create_sheet("Exceptions")
    headers = [
        "Exception ID",
        "Workflow",
        "Workbook",
        "Row",
        "Account name",
        "Account number",
        "Currency",
        "Reasons",
        "Narrative",
        "Matched counterparty",
        "Matched project code",
        "Resolved position",
        "Classification",
        "Recommended owner",
        "How to reconcile",
        "Evidence required",
        "Completion check",
    ]
    exceptions_sheet.append(headers)
    for cell in exceptions_sheet[1]:
        cell.font = Font(bold=True)

    for item in _workflow_exceptions(result):
        guidance = item.get("reconciliation_guidance", [])
        owners = sorted(
            {
                _text(section.get("owner"))
                for section in guidance
                if isinstance(section, dict) and _text(section.get("owner"))
            }
        )
        evidence = " | ".join(
            _text(section.get("evidence_required"))
            for section in guidance
            if isinstance(section, dict) and _text(section.get("evidence_required"))
        )
        completion = " | ".join(
            _text(section.get("completion_check"))
            for section in guidance
            if isinstance(section, dict) and _text(section.get("completion_check"))
        )
        exceptions_sheet.append(
            [
                item.get("exception_id", ""),
                item.get("workflow", ""),
                item.get("workbook", ""),
                item.get("row", ""),
                item.get("account_name", ""),
                item.get("account_number", ""),
                item.get("currency", ""),
                " | ".join(item.get("reasons", [])),
                item.get("narrative", ""),
                item.get("matched_counterparty", ""),
                item.get("matched_project_code", ""),
                item.get("resolved_position", ""),
                item.get("classification", ""),
                ", ".join(owners),
                _guidance_text(item),
                evidence,
                completion,
            ]
        )

    evidence_sheet = workbook.create_sheet("Evidence")
    evidence_sheet.append(["Kind", "File name", "SHA-256"])
    for cell in evidence_sheet[1]:
        cell.font = Font(bold=True)
    evidence = result.get("evidence", {})
    for kind, key in (("PDF", "pdf_sha256"), ("Excel", "excel_sha256")):
        for item in evidence.get(key, []):
            if isinstance(item, dict):
                evidence_sheet.append([kind, item.get("file_name", ""), item.get("sha256", "")])

    _autosize(workbook)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_ylookup_pdf_report(result: dict[str, Any]) -> bytes:
    try:
        from reportlab.lib import colors  # type: ignore[import-untyped]
        from reportlab.lib.pagesizes import A4  # type: ignore[import-untyped]
        from reportlab.lib.styles import getSampleStyleSheet  # type: ignore[import-untyped]
        from reportlab.lib.units import mm  # type: ignore[import-untyped]
        from reportlab.platypus import (  # type: ignore[import-untyped]
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("reportlab is required to generate PDF reports") from exc

    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title="Cherry FundOps Ylookup Review Report",
    )
    styles = getSampleStyleSheet()
    story: list[Any] = [
        Paragraph("Cherry FundOps — Ylookup Review Report", styles["Title"]),
        Paragraph(
            "Evidence-backed review only; Cherry FundOps does not initiate payments.",
            styles["BodyText"],
        ),
        Spacer(1, 8),
    ]

    input_summary = result.get("input_summary", {})
    story.append(
        Table(
            [
                ["Generated", datetime.now(UTC).isoformat()],
                ["PDF files", input_summary.get("pdf_count", 0)],
                ["Excel workbooks", input_summary.get("excel_count", 0)],
                ["Workflows", len(result.get("workflows", []))],
            ],
            colWidths=[50 * mm, 120 * mm],
        )
    )
    story.append(Spacer(1, 12))

    for workflow in result.get("workflows", []):
        if not isinstance(workflow, dict):
            continue
        story.append(Paragraph(_text(workflow.get("title", "Workflow")), styles["Heading2"]))
        story.append(
            Paragraph(
                f"Status: {_text(workflow.get('status'))} · "
                f"Workbook: {_text(workflow.get('workbook'))}",
                styles["BodyText"],
            )
        )
        metrics: list[list[Any]] = []
        for label, key in (
            ("Statement rows", "total_transactions"),
            ("Journal lines", "journal_lines"),
            ("Rows needing attention", "review_queue_rows"),
            ("GL rows", "row_count"),
            ("Investors", "investor_count"),
            ("Deals", "deal_count"),
        ):
            if key in workflow:
                metrics.append([label, workflow[key]])
        if metrics:
            story.append(Table(metrics, colWidths=[60 * mm, 30 * mm]))
            story.append(Spacer(1, 8))

        exceptions = workflow.get("exceptions") or workflow.get("sample_exceptions") or []
        if exceptions:
            story.append(Paragraph("Exception review queue", styles["Heading3"]))
        for item in exceptions:
            if not isinstance(item, dict):
                continue
            exception_id = _text(item.get("exception_id")) or f"Row {item.get('row', '')}"
            reasons = " · ".join(item.get("reasons", []))
            story.append(Paragraph(f"<b>{exception_id}</b> — {reasons}", styles["BodyText"]))
            narrative = _text(item.get("narrative"))
            if narrative:
                story.append(Paragraph(narrative, styles["BodyText"]))
            for guidance in item.get("reconciliation_guidance", []):
                if not isinstance(guidance, dict):
                    continue
                story.append(
                    Paragraph(
                        f"<b>{_text(guidance.get('reason'))}</b> · "
                        f"Owner: {_text(guidance.get('owner'))}",
                        styles["BodyText"],
                    )
                )
                for step in guidance.get("steps", []):
                    story.append(Paragraph(f"• {_text(step)}", styles["BodyText"]))
                story.append(
                    Paragraph(
                        f"Evidence: {_text(guidance.get('evidence_required'))}<br/>"
                        f"Completion: {_text(guidance.get('completion_check'))}",
                        styles["BodyText"],
                    )
                )
            story.append(Spacer(1, 8))
        story.append(PageBreak())

    story.append(Paragraph("Evidence hashes", styles["Heading2"]))
    evidence_rows = [["Kind", "File", "SHA-256"]]
    evidence = result.get("evidence", {})
    for kind, key in (("PDF", "pdf_sha256"), ("Excel", "excel_sha256")):
        for item in evidence.get(key, []):
            if isinstance(item, dict):
                evidence_rows.append(
                    [kind, _text(item.get("file_name")), _text(item.get("sha256"))]
                )
    table = Table(evidence_rows, colWidths=[16 * mm, 58 * mm, 102 * mm], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123c32")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d6d2c8")),
            ]
        )
    )
    story.append(table)
    document.build(story)
    return output.getvalue()
