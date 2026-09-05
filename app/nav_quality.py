"""NAV Quality Controller: deterministic second-line review of an administrator's NAV pack.

Problem: a fund administrator delivers a draft NAV (balance sheet, NAV bridge, investor capital
statement) that often takes several review rounds before the manager can trust it — errors surface
as a balance sheet that does not reconcile to equity, a NAV bridge that does not foot, an
independently-recomputed NAV that disagrees with the administrator's figure, or an investor capital
account that ignores a side-letter term.

Architecture, matching the rest of this codebase and the private-markets pipeline in particular:
LLM/manual entry is only ever used to produce the *inputs* (an administrator's reported NAV summary,
optionally a source ledger, optionally side-letter rules); every check below is deterministic Decimal
arithmetic. Nothing here writes to a general ledger or alters the official NAV — it only produces
findings, a recommended action and evidence for a human reviewer.

``parse_investor_level_gl_workbook`` reads a real fund-accounting export shape (confirmed against an
anonymised investor-level GL sample: 43 columns, two duplicate-named ``Static Date`` and ``GL Date``
columns, one sheet named ``Investor-Level GL``) by fixed column position rather than by header-name
lookup, because that source format repeats column names and a name-keyed dict would silently drop
the first occurrence.
"""

from __future__ import annotations

import io
import json
from datetime import date
from decimal import Decimal
from enum import StrEnum
from typing import Any, Literal

from openpyxl import load_workbook
from pydantic import BaseModel, Field, field_validator

from app.private_markets import FindingSeverity, WorkItemPriority, money

REQUIRED_GL_SHEET = "Investor-Level GL"

# Fixed column positions in the real "Investor-Level GL" export. Two column pairs repeat their
# header text ("Static Date" for period start/end, "GL Date" twice), so a header-name lookup would
# silently collapse to the last occurrence — position is the only reliable key for this format.
_GL_COLUMN = {
    "period_start": 1,
    "period_end": 2,
    "legal_entity": 3,
    "account_type": 21,
    "trans_type": 22,
    "gl_date": 23,
    "entity_currency": 30,
    "amount_entity_currency": 31,
    "investor": 35,
}
_KNOWN_ACCOUNT_TYPES = {"Assets", "Liabilities", "Capital", "Revenues", "Expenses"}


def _text(value: Any) -> str:
    return str(value or "").strip()


def _as_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip()[:10])


class NAVAction(StrEnum):
    READY_TO_SUBMIT = "ready_to_submit"
    NEEDS_REVIEW = "needs_review"
    RETURN_TO_ADMINISTRATOR = "return_to_administrator"


class NAVFinding(BaseModel):
    code: str
    severity: FindingSeverity
    title: str
    detail: str
    expected: str | None = None
    observed: str | None = None


class NAVWorkItem(BaseModel):
    code: str
    priority: WorkItemPriority
    owner: str
    title: str
    instruction: str


class NAVLedgerEntry(BaseModel):
    legal_entity: str
    account_type: str
    trans_type: str
    gl_date: date
    amount: Decimal
    currency: str = "USD"
    investor: str | None = None

    @field_validator("amount", mode="before")
    @classmethod
    def normalise_amount(cls, value: Any) -> Decimal:
        return money(value)


class NAVSourceLedger(BaseModel):
    period_start: date
    period_end: date
    entries: list[NAVLedgerEntry] = Field(default_factory=list)
    warnings: list[str] = Field(default_factory=list)

    def _entity_entries(self, legal_entity: str) -> list[NAVLedgerEntry]:
        needle = legal_entity.casefold()
        return [entry for entry in self.entries if entry.legal_entity.casefold() == needle]

    def capital_balance(
        self, legal_entity: str, *, as_of: date, investor: str | None = None
    ) -> Decimal:
        """Partners' capital (NAV) as of a date, sign-flipped to a conventional positive balance.

        The source ledger is credit-normal (assets positive, capital/liabilities/revenue negative,
        expenses positive) so that every account type for one legal entity sums to exactly zero —
        confirmed against the real anonymised export. Flipping the sign here presents NAV the way an
        administrator reports it.
        """

        total = Decimal("0")
        for entry in self._entity_entries(legal_entity):
            if entry.account_type != "Capital" or entry.gl_date > as_of:
                continue
            if investor is not None and (entry.investor or "").casefold() != investor.casefold():
                continue
            total += entry.amount
        return money(-total)

    def balance(self, legal_entity: str, account_type: str, *, as_of: date) -> Decimal:
        total = Decimal("0")
        for entry in self._entity_entries(legal_entity):
            if entry.account_type == account_type and entry.gl_date <= as_of:
                total += entry.amount
        return money(total)

    def investors(self, legal_entity: str) -> list[str]:
        seen: dict[str, None] = {}
        for entry in self._entity_entries(legal_entity):
            if entry.account_type == "Capital" and entry.investor:
                seen.setdefault(entry.investor, None)
        return list(seen)


def parse_investor_level_gl_workbook(content: bytes) -> NAVSourceLedger:
    if not content:
        raise ValueError("The source ledger workbook is empty.")
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    if REQUIRED_GL_SHEET not in workbook.sheetnames:
        raise ValueError(f"Workbook must contain an {REQUIRED_GL_SHEET!r} sheet.")

    sheet = workbook[REQUIRED_GL_SHEET]
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration as exc:
        raise ValueError(f"{REQUIRED_GL_SHEET} contains no rows.") from exc

    required_width = max(_GL_COLUMN.values()) + 1
    if len(header) < required_width:
        raise ValueError(
            f"{REQUIRED_GL_SHEET} has {len(header)} columns; expected at least {required_width}."
        )
    expected_labels = {
        "legal_entity": "Legal Entity",
        "account_type": "Account Type",
        "trans_type": "Trans Type",
        "amount_entity_currency": "Amount (Entity Currency)",
        "investor": "Investor",
    }
    for field, label in expected_labels.items():
        column = _GL_COLUMN[field]
        if _text(header[column]) != label:
            raise ValueError(
                f"{REQUIRED_GL_SHEET} column {column} is {header[column]!r}; expected {label!r}. "
                "The export layout may have changed."
            )

    entries: list[NAVLedgerEntry] = []
    warnings: list[str] = []
    period_start: date | None = None
    period_end: date | None = None
    unclassified_account_types: set[str] = set()

    for index, row in enumerate(rows, start=2):
        legal_entity = _text(row[_GL_COLUMN["legal_entity"]])
        if not legal_entity:
            continue
        account_type = _text(row[_GL_COLUMN["account_type"]])
        gl_date = _as_date(row[_GL_COLUMN["gl_date"]])
        amount = row[_GL_COLUMN["amount_entity_currency"]]
        if gl_date is None or amount is None:
            warnings.append(f"Row {index} is missing a GL date or amount and was skipped.")
            continue
        if account_type not in _KNOWN_ACCOUNT_TYPES:
            unclassified_account_types.add(account_type or "(blank)")

        if period_start is None:
            period_start = _as_date(row[_GL_COLUMN["period_start"]])
        if period_end is None:
            period_end = _as_date(row[_GL_COLUMN["period_end"]])

        entries.append(
            NAVLedgerEntry(
                legal_entity=legal_entity,
                account_type=account_type,
                trans_type=_text(row[_GL_COLUMN["trans_type"]]),
                gl_date=gl_date,
                amount=amount,
                currency=_text(row[_GL_COLUMN["entity_currency"]]) or "USD",
                investor=_text(row[_GL_COLUMN["investor"]]) or None,
            )
        )

    if not entries:
        raise ValueError(f"{REQUIRED_GL_SHEET} produced no usable rows.")
    if unclassified_account_types:
        warnings.append(
            "Unrecognised account type(s) present and excluded from balance-sheet totals: "
            + ", ".join(sorted(unclassified_account_types))
        )
    if period_start is None or period_end is None:
        raise ValueError(
            f"{REQUIRED_GL_SHEET} is missing the reporting period (Static Date columns)."
        )

    return NAVSourceLedger(
        period_start=period_start, period_end=period_end, entries=entries, warnings=warnings
    )


class InvestorCapitalLine(BaseModel):
    investor: str
    reported_capital: Decimal
    management_fee: Decimal | None = None

    @field_validator("reported_capital", "management_fee", mode="before")
    @classmethod
    def normalise_money(cls, value: Any) -> Decimal | None:
        if value is None or value == "":
            return None
        return money(value)


class AdministratorNAVSummary(BaseModel):
    legal_entity: str
    period_end: date
    currency: str = "USD"
    total_assets: Decimal
    total_liabilities: Decimal
    reported_equity: Decimal
    opening_nav: Decimal
    closing_nav: Decimal
    contributions: Decimal = Decimal("0")
    distributions: Decimal = Decimal("0")
    investment_movement: Decimal = Decimal("0")
    income: Decimal = Decimal("0")
    expenses: Decimal = Decimal("0")
    fx_movement: Decimal = Decimal("0")
    investor_capital: list[InvestorCapitalLine] = Field(default_factory=list)

    @field_validator(
        "total_assets",
        "total_liabilities",
        "reported_equity",
        "opening_nav",
        "closing_nav",
        "contributions",
        "distributions",
        "investment_movement",
        "income",
        "expenses",
        "fx_movement",
        mode="before",
    )
    @classmethod
    def normalise_money(cls, value: Any) -> Decimal:
        return money(value)

    @field_validator("currency")
    @classmethod
    def normalise_currency(cls, value: str) -> str:
        return value.upper().strip() or "USD"


class SideLetterRule(BaseModel):
    investor: str
    rule: str
    source: str | None = None
    document_id: str | None = None
    document_name: str | None = None
    section_reference: str | None = None
    page_number: int | None = Field(default=None, ge=1)
    source_excerpt: str | None = Field(default=None, max_length=1_000)
    source_sha256: str | None = Field(
        default=None,
        min_length=64,
        max_length=64,
        pattern=r"^[0-9a-f]{64}$",
    )
    effective_date: date | None = None
    explicit_override: bool = False
    resolution_status: Literal["found", "review_required", "conflict"] = "found"
    resolution_explanation: str | None = None

    def evidence_issue(self, as_of: date) -> str | None:
        if self.resolution_status != "found":
            return self.resolution_explanation or (
                "The investor-specific contract rule is unresolved."
            )
        if not self.explicit_override:
            return "The source does not contain an explicit override of the LPA default."
        if self.effective_date is None:
            return "The side-letter rule has no effective date."
        if self.effective_date > as_of:
            return "The side-letter rule is not yet effective for this reporting date."
        if not all(
            (
                self.document_id,
                self.document_name,
                self.section_reference,
                self.page_number,
                self.source_excerpt,
                self.source_sha256,
            )
        ):
            return "The side-letter rule does not have a complete document source locator."
        return None


_REQUIRED_SUMMARY_FIELDS = (
    "legal_entity",
    "period_end",
    "total_assets",
    "total_liabilities",
    "reported_equity",
    "opening_nav",
    "closing_nav",
)


def parse_administrator_nav_summary(content: bytes) -> AdministratorNAVSummary:
    """Parse the administrator's reported NAV summary (flexible JSON, like ``parse_cash_json``)."""

    if not content:
        raise ValueError("The administrator NAV summary is empty.")
    try:
        payload: dict[str, Any] = json.loads(content.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("Administrator NAV summary must be valid UTF-8 JSON.") from exc
    if not isinstance(payload, dict):
        raise ValueError("Administrator NAV summary must be a JSON object.")

    missing = [field for field in _REQUIRED_SUMMARY_FIELDS if payload.get(field) in (None, "")]
    if missing:
        raise ValueError(
            "Administrator NAV summary is missing required field(s): " + ", ".join(missing)
        )

    investor_capital_raw = payload.get("investor_capital") or []
    if not isinstance(investor_capital_raw, list):
        raise ValueError("investor_capital must be an array.")
    investor_capital = [
        InvestorCapitalLine(
            investor=_text(row.get("investor")),
            reported_capital=row.get("reported_capital"),
            management_fee=row.get("management_fee"),
        )
        for row in investor_capital_raw
    ]

    return AdministratorNAVSummary(
        legal_entity=_text(payload.get("legal_entity")),
        period_end=payload.get("period_end"),
        currency=payload.get("currency") or "USD",
        total_assets=payload.get("total_assets"),
        total_liabilities=payload.get("total_liabilities"),
        reported_equity=payload.get("reported_equity"),
        opening_nav=payload.get("opening_nav"),
        closing_nav=payload.get("closing_nav"),
        contributions=payload.get("contributions", 0),
        distributions=payload.get("distributions", 0),
        investment_movement=payload.get("investment_movement", 0),
        income=payload.get("income", 0),
        expenses=payload.get("expenses", 0),
        fx_movement=payload.get("fx_movement", 0),
        investor_capital=investor_capital,
    )


def parse_side_letter_rules(content: bytes) -> list[SideLetterRule]:
    if not content:
        return []
    try:
        payload = json.loads(content.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("Side-letter rules must be valid UTF-8 JSON.") from exc
    rows = payload.get("rules") if isinstance(payload, dict) else payload
    if not isinstance(rows, list):
        raise ValueError("Side-letter rules must be an array, or an object with a rules array.")
    return [
        SideLetterRule(
            investor=_text(row.get("investor")),
            rule=_text(row.get("rule")),
            source=_text(row.get("source")) or None,
            document_id=_text(row.get("document_id")) or None,
            document_name=_text(row.get("document_name")) or None,
            section_reference=_text(row.get("section_reference")) or None,
            page_number=row.get("page_number"),
            source_excerpt=_text(row.get("source_excerpt")) or None,
            source_sha256=_text(row.get("source_sha256")) or None,
            effective_date=row.get("effective_date"),
            explicit_override=bool(row.get("explicit_override", False)),
            resolution_status=_text(row.get("resolution_status")) or "found",
            resolution_explanation=_text(row.get("resolution_explanation")) or None,
        )
        for row in rows
    ]


class BalanceSheetCheck(BaseModel):
    reported_assets: Decimal
    reported_liabilities: Decimal
    reported_equity: Decimal
    footed_equity: Decimal
    ledger_assets: Decimal | None = None
    ledger_liabilities: Decimal | None = None


class NAVBridgeCheck(BaseModel):
    opening_nav: Decimal
    contributions: Decimal
    distributions: Decimal
    investment_movement: Decimal
    income: Decimal
    expenses: Decimal
    fx_movement: Decimal
    reported_closing_nav: Decimal
    bridge_calculated_closing_nav: Decimal
    ledger_closing_nav: Decimal | None = None


class InvestorCapitalCheck(BaseModel):
    investor: str
    reported_capital: Decimal
    ledger_capital: Decimal | None = None
    side_letter_rule: str | None = None
    rule_adjusted_expected: Decimal | None = None


class NAVReviewReport(BaseModel):
    legal_entity: str
    period_end: date
    currency: str
    action: NAVAction
    balance_sheet: BalanceSheetCheck
    nav_bridge: NAVBridgeCheck
    investor_reconciliation: list[InvestorCapitalCheck] = Field(default_factory=list)
    findings: list[NAVFinding] = Field(default_factory=list)
    work_items: list[NAVWorkItem] = Field(default_factory=list)
    controls_passed: int = 0
    exceptions_open: int = 0
    controls_summary: str
    financial_boundary: str = "Decision support only; this service never posts a journal entry or amends the official NAV."


def _side_letter_lookup(rules: list[SideLetterRule], investor: str) -> SideLetterRule | None:
    matches = [rule for rule in rules if rule.investor.casefold() == investor.casefold()]
    if len(matches) <= 1:
        return matches[0] if matches else None
    return matches[0].model_copy(
        update={
            "resolution_status": "conflict",
            "resolution_explanation": (
                "Multiple investor-specific rules match this investor; precedence must be reviewed."
            ),
        }
    )


def review_nav_quality(
    summary: AdministratorNAVSummary,
    ledger: NAVSourceLedger | None = None,
    side_letter_rules: list[SideLetterRule] | None = None,
) -> NAVReviewReport:
    """Run the deterministic NAV quality checks against an administrator's reported NAV summary."""

    rules = side_letter_rules or []
    findings: list[NAVFinding] = []
    work_items: list[NAVWorkItem] = []

    # --- Check 1: balance sheet <-> equity -----------------------------------------------------
    footed_equity = money(summary.total_assets - summary.total_liabilities)
    footing_difference = money(summary.reported_equity - footed_equity)
    if footing_difference == 0:
        findings.append(
            NAVFinding(
                code="balance_sheet.footing_valid",
                severity=FindingSeverity.PASS,
                title="Balance sheet reconciles to reported equity",
                detail="Reported assets less liabilities equals reported equity.",
                observed=str(footed_equity),
            )
        )
    else:
        findings.append(
            NAVFinding(
                code="balance_sheet.footing_mismatch",
                severity=FindingSeverity.HIGH,
                title="Balance sheet does not reconcile to reported equity",
                detail="Assets less liabilities does not equal the reported equity figure.",
                expected=str(footed_equity),
                observed=str(summary.reported_equity),
            )
        )

    ledger_assets: Decimal | None = None
    ledger_liabilities: Decimal | None = None
    ledger_closing_nav: Decimal | None = None
    if ledger is not None:
        ledger_assets = ledger.balance(summary.legal_entity, "Assets", as_of=summary.period_end)
        # The source ledger is credit-normal (liabilities negative); flip the sign so it is
        # comparable to how an administrator conventionally reports a positive liabilities total.
        ledger_liabilities = money(
            -ledger.balance(summary.legal_entity, "Liabilities", as_of=summary.period_end)
        )
        assets_diff = money(summary.total_assets - ledger_assets)
        liabilities_diff = money(summary.total_liabilities - ledger_liabilities)
        if assets_diff == 0 and liabilities_diff == 0:
            findings.append(
                NAVFinding(
                    code="balance_sheet.matches_ledger",
                    severity=FindingSeverity.PASS,
                    title="Balance sheet matches the source ledger",
                    detail="Reported assets and liabilities match the independent ledger totals.",
                )
            )
        else:
            findings.append(
                NAVFinding(
                    code="balance_sheet.ledger_mismatch",
                    severity=FindingSeverity.HIGH,
                    title="Balance sheet differs from the source ledger",
                    detail=(
                        "The administrator's reported assets/liabilities do not match the "
                        "independently derived ledger totals for this entity and period."
                    ),
                    expected=f"assets {ledger_assets}, liabilities {ledger_liabilities}",
                    observed=f"assets {summary.total_assets}, liabilities {summary.total_liabilities}",
                )
            )

    # --- Check 2: NAV bridge --------------------------------------------------------------------
    bridge_calculated_closing = money(
        summary.opening_nav
        + summary.contributions
        + summary.investment_movement
        + summary.income
        + summary.fx_movement
        - summary.expenses
        - summary.distributions
    )
    bridge_difference = money(summary.closing_nav - bridge_calculated_closing)
    if bridge_difference == 0:
        findings.append(
            NAVFinding(
                code="nav_bridge.foots",
                severity=FindingSeverity.PASS,
                title="NAV bridge foots",
                detail="Opening NAV plus reported movements equals the reported closing NAV.",
                observed=str(bridge_calculated_closing),
            )
        )
    else:
        findings.append(
            NAVFinding(
                code="nav_bridge.does_not_foot",
                severity=FindingSeverity.HIGH,
                title="NAV bridge does not foot",
                detail="Opening NAV plus reported movements does not equal the reported closing NAV.",
                expected=str(bridge_calculated_closing),
                observed=str(summary.closing_nav),
            )
        )

    if ledger is not None:
        ledger_closing_nav = ledger.capital_balance(summary.legal_entity, as_of=summary.period_end)
        recalculation_difference = money(summary.closing_nav - ledger_closing_nav)
        if recalculation_difference == 0:
            findings.append(
                NAVFinding(
                    code="nav.independent_recalculation_valid",
                    severity=FindingSeverity.PASS,
                    title="Independent NAV recalculation matches",
                    detail="The reported closing NAV matches the NAV independently derived from the ledger.",
                )
            )
        else:
            findings.append(
                NAVFinding(
                    code="nav.independent_recalculation_mismatch",
                    severity=FindingSeverity.HIGH,
                    title="Independent NAV recalculation disagrees with the reported NAV",
                    detail=(
                        "Recomputing partners' capital directly from the source ledger produces a "
                        "different closing NAV than the administrator reported."
                    ),
                    expected=str(ledger_closing_nav),
                    observed=str(summary.closing_nav),
                )
            )

    # --- Check 3: investor capital reconciliation -----------------------------------------------
    investor_checks: list[InvestorCapitalCheck] = []
    for line in summary.investor_capital:
        ledger_capital = (
            ledger.capital_balance(
                summary.legal_entity, as_of=summary.period_end, investor=line.investor
            )
            if ledger is not None
            else None
        )
        rule = _side_letter_lookup(rules, line.investor)
        rule_adjusted_expected: Decimal | None = None
        check = InvestorCapitalCheck(
            investor=line.investor,
            reported_capital=line.reported_capital,
            ledger_capital=ledger_capital,
            side_letter_rule=rule.rule if rule else None,
        )

        evidence_issue = rule.evidence_issue(summary.period_end) if rule is not None else None
        if rule is not None and evidence_issue:
            findings.append(
                NAVFinding(
                    code="side_letter.evidence_incomplete",
                    severity=FindingSeverity.WARNING,
                    title=f"{line.investor}: contract evidence requires review",
                    detail=(
                        f"The investor-specific rule was not applied automatically. {evidence_issue}"
                    ),
                )
            )
        elif rule is not None and rule.rule != "management_fee_offsets_called_capital":
            findings.append(
                NAVFinding(
                    code="side_letter.rule_unsupported",
                    severity=FindingSeverity.WARNING,
                    title=f"{line.investor}: unsupported side-letter rule",
                    detail=(
                        "The source-backed rule uses calculation semantics that this NAV control "
                        "does not support, so it requires human review."
                    ),
                )
            )
        elif rule is not None:
            if line.management_fee is None:
                findings.append(
                    NAVFinding(
                        code="side_letter.missing_management_fee",
                        severity=FindingSeverity.WARNING,
                        title=f"{line.investor}: side-letter rule cannot be validated",
                        detail=(
                            f"Side Letter ({rule.source or 'on file'}) requires the management fee to "
                            "offset called capital, but no management fee was supplied for this investor."
                        ),
                    )
                )
            else:
                baseline = ledger_capital if ledger_capital is not None else line.reported_capital
                rule_adjusted_expected = money(baseline - line.management_fee)
                check.rule_adjusted_expected = rule_adjusted_expected
                rule_difference = money(line.reported_capital - rule_adjusted_expected)
                if rule_difference == 0:
                    findings.append(
                        NAVFinding(
                            code="side_letter.rule_applied",
                            severity=FindingSeverity.PASS,
                            title=f"{line.investor}: side-letter rule correctly applied",
                            detail=(
                                f"Management fee correctly offsets called capital per "
                                f"{rule.source or 'the side letter'}."
                            ),
                        )
                    )
                else:
                    findings.append(
                        NAVFinding(
                            code="side_letter.rule_violation",
                            severity=FindingSeverity.HIGH,
                            title=f"{line.investor}: side-letter rule not applied",
                            detail=(
                                f"Management fee of {line.management_fee} must offset called capital "
                                f"per {rule.source or 'the side letter'}, but the reported capital "
                                "does not reflect this."
                            ),
                            expected=str(rule_adjusted_expected),
                            observed=str(line.reported_capital),
                        )
                    )
        elif ledger_capital is not None:
            capital_difference = money(line.reported_capital - ledger_capital)
            if capital_difference == 0:
                findings.append(
                    NAVFinding(
                        code="investor_capital.matches_ledger",
                        severity=FindingSeverity.PASS,
                        title=f"{line.investor}: capital account matches the ledger",
                        detail="Reported investor capital matches the independently derived ledger balance.",
                    )
                )
            else:
                findings.append(
                    NAVFinding(
                        code="investor_capital.ledger_mismatch",
                        severity=FindingSeverity.HIGH,
                        title=f"{line.investor}: capital account differs from the ledger",
                        detail=(
                            "Reported investor capital does not match the independently derived "
                            "ledger balance for this investor."
                        ),
                        expected=str(ledger_capital),
                        observed=str(line.reported_capital),
                    )
                )
        investor_checks.append(check)

    high_codes = {finding.code for finding in findings if finding.severity == FindingSeverity.HIGH}
    warning_codes = {
        finding.code for finding in findings if finding.severity == FindingSeverity.WARNING
    }

    if high_codes:
        action = NAVAction.RETURN_TO_ADMINISTRATOR
    elif warning_codes:
        action = NAVAction.NEEDS_REVIEW
    else:
        action = NAVAction.READY_TO_SUBMIT

    if (
        "balance_sheet.footing_mismatch" in high_codes
        or "balance_sheet.ledger_mismatch" in high_codes
    ):
        work_items.append(
            NAVWorkItem(
                code="resolve_balance_sheet",
                priority=WorkItemPriority.CRITICAL,
                owner="Fund controller",
                title="Resolve the balance sheet break",
                instruction="Trace assets, liabilities and equity to source and correct the NAV pack.",
            )
        )
    if (
        "nav_bridge.does_not_foot" in high_codes
        or "nav.independent_recalculation_mismatch" in high_codes
    ):
        work_items.append(
            NAVWorkItem(
                code="resolve_nav_bridge",
                priority=WorkItemPriority.CRITICAL,
                owner="Fund controller",
                title="Resolve the NAV bridge break",
                instruction=(
                    "Reconcile opening NAV, movements and closing NAV against the source ledger "
                    "before returning the pack to the administrator."
                ),
            )
        )
    if high_codes & {"investor_capital.ledger_mismatch", "side_letter.rule_violation"}:
        work_items.append(
            NAVWorkItem(
                code="resolve_investor_capital",
                priority=WorkItemPriority.HIGH,
                owner="Investor relations",
                title="Resolve investor capital account discrepancies",
                instruction=(
                    "Correct each flagged investor's capital account, applying any side-letter terms, "
                    "before the NAV is released."
                ),
            )
        )
    if "side_letter.missing_management_fee" in warning_codes:
        work_items.append(
            NAVWorkItem(
                code="obtain_management_fee",
                priority=WorkItemPriority.NORMAL,
                owner="Investor relations",
                title="Obtain the missing management fee figure",
                instruction="Request the investor's management fee so the side-letter rule can be checked.",
            )
        )
    if warning_codes & {"side_letter.evidence_incomplete", "side_letter.rule_unsupported"}:
        work_items.append(
            NAVWorkItem(
                code="review_contract_evidence",
                priority=WorkItemPriority.HIGH,
                owner="Fund controller",
                title="Review investor-specific contract evidence",
                instruction=(
                    "Confirm the investor identity, explicit override, effective date and exact "
                    "document locator before applying the term to the NAV calculation."
                ),
            )
        )

    controls_passed = sum(finding.severity == FindingSeverity.PASS for finding in findings)
    exceptions_open = sum(
        finding.severity in {FindingSeverity.HIGH, FindingSeverity.WARNING} for finding in findings
    )
    controls_summary = (
        f"{controls_passed} controls passed; {len(high_codes)} critical exception(s) found."
        if high_codes
        else f"{controls_passed} controls passed; {len(warning_codes)} item(s) need review."
        if warning_codes
        else f"{controls_passed} controls passed; ready to submit."
    )

    return NAVReviewReport(
        legal_entity=summary.legal_entity,
        period_end=summary.period_end,
        currency=summary.currency,
        action=action,
        balance_sheet=BalanceSheetCheck(
            reported_assets=summary.total_assets,
            reported_liabilities=summary.total_liabilities,
            reported_equity=summary.reported_equity,
            footed_equity=footed_equity,
            ledger_assets=ledger_assets,
            ledger_liabilities=ledger_liabilities,
        ),
        nav_bridge=NAVBridgeCheck(
            opening_nav=summary.opening_nav,
            contributions=summary.contributions,
            distributions=summary.distributions,
            investment_movement=summary.investment_movement,
            income=summary.income,
            expenses=summary.expenses,
            fx_movement=summary.fx_movement,
            reported_closing_nav=summary.closing_nav,
            bridge_calculated_closing_nav=bridge_calculated_closing,
            ledger_closing_nav=ledger_closing_nav,
        ),
        investor_reconciliation=investor_checks,
        findings=findings,
        work_items=work_items,
        controls_passed=controls_passed,
        exceptions_open=exceptions_open,
        controls_summary=controls_summary,
    )
