"""Gemini-based extraction of an administrator NAV summary from an arbitrary NAV pack (PDF or
Excel), so a fund manager can upload the pack itself rather than a pre-built nav_summary.json.

Same boundary as app.document_ai.GeminiDocumentExtractor: Gemini only extracts what is visible in
the document into the same AdministratorNAVSummary schema app.nav_quality already validates and
review_nav_quality already checks -- it never decides whether the NAV is correct. That remains
entirely deterministic Decimal arithmetic downstream, exactly as before this module existed.

Excel workbooks are rendered to a bounded plain-text sheet dump with openpyxl before being sent to
Gemini as text, rather than uploaded as a raw binary part -- this keeps the same
"deterministic-text-extraction before any LLM step" boundary used elsewhere in this codebase (e.g.
app.contracts.read_document_pages), and avoids depending on Gemini's binary-XLSX file support.
"""

from __future__ import annotations

import io
import json
import logging

from openpyxl import load_workbook

from app.config import Settings
from app.document_ai import GeminiUnavailable
from app.nav_quality import AdministratorNAVSummary

logger = logging.getLogger(__name__)

MAX_SHEET_ROWS = 200

EXTRACTION_INSTRUCTION = """
You are the NAV-summary extraction specialist inside Cherry Agent's NAV Guardian. Extract only the
administrator's reported NAV summary figures visible in the supplied fund NAV pack: the legal
entity name, reporting period end, the balance sheet (total assets, total liabilities, reported
equity), the NAV bridge (opening NAV, contributions, distributions, investment movement, income,
expenses, FX movement, closing NAV) and each investor's reported capital (and management fee, if
stated). Never invent a figure that is not visible in the document. Use 0 for a NAV-bridge line
item that is genuinely absent from the document (for example, no FX movement occurred this
period), but do not guess a required top-line balance-sheet or NAV figure that is not stated. This
extraction never decides whether the NAV is correct; deterministic checks do that separately.
""".strip()


def _render_workbook_as_text(content: bytes, max_rows: int = MAX_SHEET_ROWS) -> str:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sections: list[str] = []
        for sheet in workbook.worksheets:
            lines = [f"Sheet: {sheet.title}"]
            truncated = False
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index >= max_rows:
                    truncated = True
                    break
                if any(value is not None for value in row):
                    lines.append(" | ".join("" if value is None else str(value) for value in row))
            if truncated:
                lines.append(f"... (truncated after {max_rows} rows)")
            sections.append("\n".join(lines))
        return "\n\n".join(sections)
    finally:
        workbook.close()


class NAVSummaryExtractor:
    def __init__(self, settings: Settings) -> None:
        self._settings = settings

    async def extract(
        self, content: bytes, mime_type: str, filename: str
    ) -> AdministratorNAVSummary:
        if not self._settings.google_ready:
            raise GeminiUnavailable(
                "Gemini is not configured. Set GOOGLE_CLOUD_PROJECT for Vertex AI "
                "or GOOGLE_API_KEY."
            )
        if not content:
            raise ValueError("The uploaded NAV pack is empty.")

        try:
            from google import genai
            from google.genai import types
        except ImportError as exc:  # pragma: no cover - exercised only in a reduced install
            raise GeminiUnavailable("Install google-genai to process real documents.") from exc

        is_pdf = mime_type == "application/pdf" or filename.casefold().endswith(".pdf")
        if is_pdf:
            parts = [
                types.Part.from_text(
                    text=(
                        f"Extract the NAV summary from {filename!r}. Return only the structured "
                        "response required by the schema."
                    )
                ),
                types.Part.from_bytes(data=content, mime_type=mime_type),
            ]
        else:
            workbook_text = _render_workbook_as_text(content)
            parts = [
                types.Part.from_text(
                    text=(
                        f"Extract the NAV summary from the workbook {filename!r} below. Return "
                        "only the structured response required by the schema.\n\n" + workbook_text
                    )
                )
            ]

        self._settings.configure_google_environment()
        client = genai.Client()
        async_client = client.aio
        try:
            response = await async_client.models.generate_content(
                model=self._settings.gemini_model,
                contents=parts,
                config=types.GenerateContentConfig(
                    system_instruction=EXTRACTION_INSTRUCTION,
                    response_mime_type="application/json",
                    response_json_schema=AdministratorNAVSummary.model_json_schema(),
                ),
            )
        finally:
            await async_client.aclose()

        if not response.text:
            raise ValueError("Gemini returned no NAV summary extraction.")
        try:
            payload = json.loads(response.text)
            return AdministratorNAVSummary.model_validate(payload)
        except (json.JSONDecodeError, ValueError) as exc:
            logger.exception("Gemini returned an invalid NAV summary extraction")
            raise ValueError("Gemini returned an invalid NAV summary extraction.") from exc
