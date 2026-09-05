"""Atomic deterministic primitives for the NAV Guardian Reconciliation Agent.

These are intentionally low-level building blocks — read a workbook, read a cell, sum values,
compare two values, build an accounting bridge — so the agent can compose an ad hoc
reconciliation instead of being limited to the fixed NAV Guardian checks in
app.nav_reconciliation. The agent decides which tool to call and why; these tools perform the
actual operation.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.nav_reconciliation import money


def read_excel(
    workbook_path: str,
    sheet_name: str | None = None,
    max_rows: int = 50,
) -> dict[str, Any]:
    """Read a worksheet's header row and up to max_rows of data rows.

    Args:
        workbook_path: Local path to an XLSX workbook.
        sheet_name: Sheet to read; defaults to the workbook's first sheet.
        max_rows: Maximum number of data rows to return, to keep tool output bounded.
    """

    path = Path(workbook_path)
    if not path.is_file():
        raise ValueError(f"Workbook path {workbook_path!r} does not exist.")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        rows_iter = sheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else "" for value in next(rows_iter, ())]

        rows: list[list[Any]] = []
        for row in rows_iter:
            if len(rows) >= max_rows:
                break
            rows.append(list(row))

        return {
            "workbook": path.name,
            "sheet_name": sheet.title,
            "sheet_names": workbook.sheetnames,
            "headers": headers,
            "row_count_returned": len(rows),
            "rows": rows,
        }
    finally:
        workbook.close()


def read_cell(workbook_path: str, sheet_name: str, cell: str) -> dict[str, Any]:
    """Read a single cell's value by its A1 reference.

    Args:
        workbook_path: Local path to an XLSX workbook.
        sheet_name: Sheet containing the cell.
        cell: A1-style cell reference, e.g. "B12".
    """

    path = Path(workbook_path)
    if not path.is_file():
        raise ValueError(f"Workbook path {workbook_path!r} does not exist.")

    workbook = load_workbook(path, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name!r} not found in {path.name!r}.")
        value = workbook[sheet_name][cell].value
        return {"workbook": path.name, "sheet_name": sheet_name, "cell": cell, "value": value}
    finally:
        workbook.close()


def calculate_sum(values: list[str]) -> dict[str, Any]:
    """Sum a list of decimal-string monetary values.

    Args:
        values: Amounts to sum, each as a decimal string, e.g. ["105200000", "-12700000"].
    """

    amounts = [money(value) for value in values]
    total = money(sum(amounts, Decimal("0")))
    return {"values": [str(amount) for amount in amounts], "total": str(total)}


def compare_values(expected: str, actual: str, tolerance: str = "0.01") -> dict[str, Any]:
    """Compare two decimal-string values within a tolerance.

    Args:
        expected: The expected/calculated value.
        actual: The reported/administrator value.
        tolerance: Maximum absolute difference still treated as a match.
    """

    expected_amount = money(expected)
    actual_amount = money(actual)
    tolerance_amount = money(tolerance)
    difference = money(expected_amount - actual_amount)
    matches = abs(difference) <= tolerance_amount
    return {
        "expected": str(expected_amount),
        "actual": str(actual_amount),
        "difference": str(difference),
        "matches": matches,
        "status": "PASS" if matches else "FAIL",
    }


def build_bridge(
    opening_balance: str,
    movements: dict[str, str],
    reported_closing: str,
    tolerance: str = "0.01",
) -> dict[str, Any]:
    """Build an ad hoc accounting bridge: opening balance plus signed movements, compared to a
    reported closing balance. Use this for bridges that don't match the fixed NAV bridge in
    validate_nav_bridge — for example a single investor's capital account.

    Args:
        opening_balance: The period's opening balance.
        movements: Signed movement amounts keyed by label, e.g. {"contributions": "500000",
            "management_fee": "-125000"}.
        reported_closing: The reported/administrator closing balance to check against.
        tolerance: Maximum absolute difference still treated as a rounding pass.
    """

    opening_amount = money(opening_balance)
    movement_amounts = {label: money(value) for label, value in movements.items()}
    reported_amount = money(reported_closing)
    tolerance_amount = money(tolerance)
    expected_amount = money(opening_amount + sum(movement_amounts.values(), Decimal("0")))
    difference = money(expected_amount - reported_amount)
    passed = abs(difference) <= tolerance_amount

    return {
        "opening_balance": str(opening_amount),
        "movements": {label: str(amount) for label, amount in movement_amounts.items()},
        "expected_closing": str(expected_amount),
        "reported_closing": str(reported_amount),
        "difference": str(difference),
        "status": "PASS" if passed else "FAIL",
    }
