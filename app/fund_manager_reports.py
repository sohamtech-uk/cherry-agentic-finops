from __future__ import annotations

import io
from html import escape
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.fund_manager_cases import FundManagerCase


def _text(value: Any) -> str:
    if value is None or value == "":
        return "—"
    return str(value)


def _items(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, dict)]


def _autosize(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, column in enumerate(sheet.columns, start=1):
            width = min(
                60,
                max(12, max((len(_text(cell.value)) for cell in column), default=0) + 2),
            )
            sheet.column_dimensions[get_column_letter(index)].width = width


def _append_header(sheet: Any, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def build_fund_manager_excel_report(case: FundManagerCase) -> bytes:
    """Build the final human-decision review workbook without returning uploaded source bytes."""

    workbook = Workbook()
    summary = workbook.active
    if summary is None:  # pragma: no cover - a new workbook always has a worksheet
        raise RuntimeError("The report workbook could not create its summary sheet.")
    summary.title = "Summary"
    summary.append(["Cherry FundOps — Fund Manager Review Report"])
    summary["A1"].font = Font(bold=True, size=16)
    summary.append(["Case ID", case.case_id])
    summary.append(["Stage", case.stage])
    summary.append(["Fund", _text(case.fund_name)])
    summary.append(["Reporting period", _text(case.reporting_period)])
    summary.append(["As-of date", _text(case.as_of_date)])
    summary.append(["Created", case.created_at])
    summary.append(["Updated", case.updated_at])
    summary.append([])
    summary.append(["Final human decision"])
    summary["A10"].font = Font(bold=True)
    decision = case.decision or {}
    summary.append(["Action", _text(decision.get("action"))])
    summary.append(["Decision note", _text(decision.get("note"))])
    summary.append(["Recorded at", _text(decision.get("recorded_at"))])
    summary.append(["Actor", _text(decision.get("actor"))])
    summary.append([])
    execution = case.execution or {}
    summary.append(["Control result", _text(execution.get("status"))])
    summary.append(["Issues found", execution.get("issues_found", 0)])
    summary.append(["Material", execution.get("material", 0)])
    summary.append(["Critical", execution.get("critical", 0)])
    summary.append([])
    summary.append(
        [
            "Control boundary",
            "AI interprets and investigates. Deterministic controls produce calculations and "
            "reconciliations. The recorded final decision is human.",
        ]
    )

    evidence_sheet = workbook.create_sheet("Evidence")
    _append_header(evidence_sheet, ["Source ID", "File", "Detected type", "Validation status"])
    for source in _items(case.classification.get("sources", [])):
        evidence_sheet.append(
            [
                source.get("id", ""),
                source.get("filename", ""),
                source.get("detected_type", ""),
                source.get("validation_status", ""),
            ]
        )

    plan_sheet = workbook.create_sheet("Control Plan")
    _append_header(plan_sheet, ["Control", "Status", "Source IDs", "Reasoning"])
    for item in _items((case.plan or {}).get("control_plan", [])):
        plan_sheet.append(
            [
                item.get("control", ""),
                item.get("status", ""),
                ", ".join(str(value) for value in item.get("source_ids", [])),
                item.get("reasoning", ""),
            ]
        )

    issues_sheet = workbook.create_sheet("Control Issues")
    _append_header(
        issues_sheet,
        ["Issue ID", "Title", "Severity", "Summary", "Recommended action"],
    )
    for issue in _items(execution.get("issues", [])):
        issues_sheet.append(
            [
                issue.get("id", ""),
                issue.get("title", ""),
                issue.get("severity", ""),
                issue.get("summary", ""),
                issue.get("recommended_action", ""),
            ]
        )

    findings_sheet = workbook.create_sheet("Findings Review")
    _append_header(
        findings_sheet,
        [
            "Issue ID",
            "Priority",
            "Finding",
            "Likely cause",
            "Evidence gap",
            "Recommended action",
        ],
    )
    investigation = case.investigation or {}
    for item in _items(investigation.get("investigations", [])):
        findings_sheet.append(
            [
                item.get("issue_id", ""),
                item.get("priority", ""),
                item.get("finding", ""),
                item.get("likely_cause", ""),
                item.get("evidence_gap", ""),
                item.get("recommended_action", ""),
            ]
        )

    _autosize(workbook)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_fund_manager_pdf_report(case: FundManagerCase) -> bytes:
    """Build a concise final PDF report for a decided Fund Manager case."""

    try:
        from reportlab.lib import colors  # type: ignore[import-untyped]
        from reportlab.lib.pagesizes import A4  # type: ignore[import-untyped]
        from reportlab.lib.styles import getSampleStyleSheet  # type: ignore[import-untyped]
        from reportlab.lib.units import mm  # type: ignore[import-untyped]
        from reportlab.platypus import (  # type: ignore[import-untyped]
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("reportlab is required to generate PDF reports") from exc

    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title="Cherry FundOps Fund Manager Review Report",
    )
    styles = getSampleStyleSheet()
    decision = case.decision or {}
    execution = case.execution or {}
    investigation = case.investigation or {}

    story: list[Any] = [
        Paragraph("Cherry FundOps — Fund Manager Review Report", styles["Title"]),
        Paragraph(
            "Final review pack after the explicit human decision. Cherry FundOps does not initiate "
            "payments or silently post accounting entries.",
            styles["BodyText"],
        ),
        Spacer(1, 8),
    ]

    metadata = [
        ["Case ID", case.case_id],
        ["Fund", _text(case.fund_name)],
        ["Reporting period", _text(case.reporting_period)],
        ["As-of date", _text(case.as_of_date)],
        ["Control result", _text(execution.get("status"))],
        ["Final decision", _text(decision.get("action")).replace("_", " ")],
        ["Decision note", _text(decision.get("note"))],
        ["Recorded at", _text(decision.get("recorded_at"))],
    ]
    metadata_table = Table(metadata, colWidths=[42 * mm, 140 * mm])
    metadata_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D8DED9")),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EFF5F0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.extend([metadata_table, Spacer(1, 12)])

    story.append(Paragraph("Evidence reviewed", styles["Heading2"]))
    evidence_rows: list[list[Any]] = [["File", "Type", "Status"]]
    for source in _items(case.classification.get("sources", [])):
        evidence_rows.append(
            [
                Paragraph(escape(_text(source.get("filename"))), styles["BodyText"]),
                Paragraph(escape(_text(source.get("detected_type"))), styles["BodyText"]),
                Paragraph(escape(_text(source.get("validation_status"))), styles["BodyText"]),
            ]
        )
    if len(evidence_rows) == 1:
        evidence_rows.append(["—", "—", "—"])
    evidence_table = Table(evidence_rows, colWidths=[86 * mm, 54 * mm, 42 * mm], repeatRows=1)
    evidence_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123C32")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D8DED9")),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.extend([evidence_table, Spacer(1, 12)])

    story.append(Paragraph("Control result", styles["Heading2"]))
    story.append(
        Paragraph(
            escape(
                f"Status: {_text(execution.get('status'))}. "
                f"Issues: {execution.get('issues_found', 0)}; "
                f"material: {execution.get('material', 0)}; "
                f"critical: {execution.get('critical', 0)}."
            ),
            styles["BodyText"],
        )
    )
    story.append(Spacer(1, 6))

    for issue in _items(execution.get("issues", [])):
        title = escape(_text(issue.get("title")))
        severity = escape(_text(issue.get("severity")))
        story.append(Paragraph(f"<b>{title}</b> — {severity}", styles["BodyText"]))
        story.append(Paragraph(escape(_text(issue.get("summary"))), styles["BodyText"]))
        story.append(
            Paragraph(
                "<b>Recommended action:</b> " + escape(_text(issue.get("recommended_action"))),
                styles["BodyText"],
            )
        )
        story.append(Spacer(1, 5))

    findings = _items(investigation.get("investigations", []))
    if findings:
        story.extend([Spacer(1, 6), Paragraph("Findings review", styles["Heading2"])])
        for item in findings:
            story.append(
                Paragraph(
                    f"<b>{escape(_text(item.get('issue_id')))}</b> — "
                    f"{escape(_text(item.get('finding')))}",
                    styles["BodyText"],
                )
            )
            if item.get("likely_cause"):
                story.append(
                    Paragraph(
                        "<b>Likely cause:</b> " + escape(_text(item.get("likely_cause"))),
                        styles["BodyText"],
                    )
                )
            if item.get("evidence_gap"):
                story.append(
                    Paragraph(
                        "<b>Evidence gap:</b> " + escape(_text(item.get("evidence_gap"))),
                        styles["BodyText"],
                    )
                )
            story.append(Spacer(1, 5))

    story.extend(
        [
            Spacer(1, 10),
            Paragraph("Human control boundary", styles["Heading2"]),
            Paragraph(
                "AI may interpret evidence, plan governed actions and investigate exceptions. "
                "Deterministic tools remain authoritative for calculations and reconciliations; "
                "the final decision shown in this report was explicitly recorded by a human user.",
                styles["BodyText"],
            ),
        ]
    )

    document.build(story)
    return output.getvalue()
