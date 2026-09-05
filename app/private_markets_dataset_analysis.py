from __future__ import annotations

import io
import os
import re
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

from app.ylookup_datasets import analyse_ylookup_dataset_batch

_PRIMARY_ACCOUNT_RE = re.compile(
    r"Account\s+number\s*[:\r\n ]+([0-9]{3}[- ]?[0-9]{6}[- ]?[0-9]{3})",
    re.IGNORECASE,
)
_ACCOUNT_NUMBER_RE = re.compile(r"\b[0-9]{3}[- ]?[0-9]{6}[- ]?[0-9]{3}\b")
_FILENAME_CODE_RE = re.compile(r"_([0-9]{4,6})(?:\([0-9]+\))?$", re.IGNORECASE)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _normalise_account_number(value: str | None) -> str | None:
    if not value:
        return None
    digits = "".join(character for character in value if character.isdigit())
    return digits if len(digits) >= 10 else None


def _extract_primary_account_number(pdf_content: bytes) -> str | None:
    if not pdf_content:
        return None
    try:
        reader = PdfReader(io.BytesIO(pdf_content), strict=False)
        text = "\n".join((page.extract_text() or "") for page in reader.pages[:2])
    except Exception:
        return None

    match = _PRIMARY_ACCOUNT_RE.search(text)
    if match:
        return _normalise_account_number(match.group(1))

    # Fallback for PDFs whose text extractor loses the label/value line break.
    early_text = text[:2500]
    fallback = _ACCOUNT_NUMBER_RE.search(early_text)
    return _normalise_account_number(fallback.group(0)) if fallback else None


def _filename_account_code(file_name: str) -> str | None:
    stem = os.path.splitext(os.path.basename(file_name))[0]
    match = _FILENAME_CODE_RE.search(stem)
    return match.group(1) if match else None


def _account_map_identifiers(workbook_content: bytes) -> tuple[set[str], set[str]]:
    account_numbers: set[str] = set()
    account_codes: set[str] = set()

    workbook = load_workbook(io.BytesIO(workbook_content), read_only=True, data_only=True)
    try:
        account_map = None
        for sheet_name in workbook.sheetnames:
            if sheet_name.strip() == "Account Map":
                account_map = workbook[sheet_name]
                break
        if account_map is None:
            return account_numbers, account_codes

        rows = account_map.iter_rows(values_only=True)
        try:
            headers = [_text(value) for value in next(rows)]
        except StopIteration:
            return account_numbers, account_codes
        indexes = {label: index for index, label in enumerate(headers) if label}

        account_number_index = indexes.get("Account Number")
        bank_account_index = indexes.get("Bank Account")
        for row in rows:
            if account_number_index is not None and account_number_index < len(row):
                normalised = _normalise_account_number(_text(row[account_number_index]))
                if normalised:
                    account_numbers.add(normalised)
            if bank_account_index is not None and bank_account_index < len(row):
                bank_account = _text(row[bank_account_index])
                match = re.search(r"([0-9]{4,6})\s*$", bank_account)
                if match:
                    account_codes.add(match.group(1))
    finally:
        workbook.close()

    return account_numbers, account_codes


def _match_bank_statement_files(
    workbook_content: bytes,
    pdf_items: list[tuple[str, bytes]],
) -> list[dict[str, str]]:
    account_numbers, account_codes = _account_map_identifiers(workbook_content)
    matches: list[dict[str, str]] = []

    for file_name, pdf_content in pdf_items:
        account_number = _extract_primary_account_number(pdf_content)
        if account_number and account_number in account_numbers:
            matches.append(
                {
                    "file_name": file_name,
                    "match_method": "statement_account_number",
                    "account_number": account_number,
                }
            )
            continue

        # Keep a filename fallback for scanned PDFs or extraction failures. Copy suffixes such as
        # "4319(1).pdf" are ignored so ordinary browser/download renaming cannot break matching.
        account_code = _filename_account_code(file_name)
        if account_code and account_code in account_codes:
            matches.append(
                {
                    "file_name": file_name,
                    "match_method": "filename_account_code",
                    "account_code": account_code,
                }
            )

    return matches


def analyse_private_markets_dataset_batch(
    workbook_items: list[tuple[str, bytes]],
    pdf_items: list[tuple[str, bytes]],
) -> dict[str, Any] | None:
    """Run the existing native-workflow analysis with product-neutral labels and PDF matching."""

    pdf_file_names = [file_name for file_name, _ in pdf_items]
    result = analyse_ylookup_dataset_batch(workbook_items, pdf_file_names)
    if result is None:
        return None

    purpose_labels = {
        "bank_statement_working": "Bank statements to journal entries working file",
        "investor_gl": "Investor-level source general ledger",
        "loader_sample": "Target-system loader sample",
        "loader_workbook": "Loader or mapping workbook",
    }
    for profile in result.get("workbook_profiles", []):
        kind = profile.get("kind")
        if kind in purpose_labels:
            profile["purpose"] = purpose_labels[kind]

    workbook_by_name = {file_name: content for file_name, content in workbook_items}
    for workflow in result.get("workflows", []):
        workflow_type = workflow.get("workflow")
        if workflow_type == "bank_statements_to_journal_entries":
            workbook_content = workbook_by_name.get(str(workflow.get("workbook")))
            if workbook_content is not None:
                statement_matches = _match_bank_statement_files(workbook_content, pdf_items)
                workflow["matched_statement_files"] = len(statement_matches)
                workflow["statement_file_matches"] = statement_matches
            workflow["message"] = (
                "The uploaded bank-statement working file was accepted directly. Review gaps are "
                "surfaced as workflow exceptions rather than treated as invalid input. Click an "
                "exception to see its reconciliation steps and required evidence."
            )
        elif workflow_type == "investor_gl_to_loader":
            if workflow.get("loader_sample_supplied"):
                workflow["message"] = (
                    "The source GL and target loader contract are both present. Mapping gaps are "
                    "surfaced explicitly rather than silently filled."
                )
            else:
                workflow["message"] = (
                    "The investor-level source GL was recognised and profiled. Add a loader sample "
                    "or mapping workbook to complete the source-to-target transformation."
                )

    result["message"] = (
        "Uploaded private-markets workbooks were auto-detected and routed to their native "
        "workflows. Bank-statement and investor-GL datasets do not require an LP_Commitments sheet."
    )
    return result
